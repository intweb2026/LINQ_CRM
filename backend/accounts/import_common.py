"""
accounts/import_common.py
──────────────────────────
Shared plumbing for the two-phase JSON-row importers.

WHY THIS FILE EXISTS
proposal_submission/importer.py was written first and paper_review/importer.py
needs the same date parsing, the same numeric/text coercion, the same header
cleaning and the same plan-hash/summary shape. Two ways to get that were
available and both were wrong:

  * COPY it into paper_review — the dirty-date fix (whitespace hugging a hyphen
    in "20 - Dec - 2025") would then need making twice, and the second copy would
    drift the first time only one was touched.
  * IMPORT from proposal_submission — a lateral dependency between two sibling
    pipeline apps whose only intended relationship is the one-directional FK
    ProposalSubmission.source_paper_review → PaperReview.

So the genuinely generic half is extracted here, exactly as
webhooks/event_code_normalization.py was extracted for spacing-tolerant event
codes. accounts/ is already this codebase's home for cross-module machinery —
filter_spec.py, bulk_update.py and ordering.py are all shared from here and used
by four or more apps.

WHAT STAYS IN EACH APP'S OWN importer.py
Everything model-specific: the Zoho header map, the reverse label map, which
columns are MR-restricted, which fields are required, and classify_rows() itself
(the classification rules differ — paper_review reconciles a stored score against
its six criteria, proposal_submission has no such rule).

accounts/bulk_update.py is NOT touched by this file; the two solve different
problems and share nothing.
"""
import hashlib
import json
import re
from datetime import date, datetime, timedelta
from html import unescape

# Per-call row cap. The browser chunks anything larger, because
# DATA_UPLOAD_MAX_MEMORY_SIZE sits at Django's 2.5 MB default and an uncapped
# paste fails opaquely rather than with a message naming the limit.
MAX_ROWS = 500

# Row classifications, shared so the frontend's three-way pill rendering is the
# same component for every importer.
CREATE = "CREATE"
CREATE_WITH_WARNING = "CREATE_WITH_WARNING"
ERROR = "ERROR"

# ── Excel serial dates ───────────────────────────────────────────────────────
# Excel's 1900 epoch, with its phantom 29-Feb-1900. Serial 1 = 1900-01-01;
# serials 1-59 are offset from 1899-12-31, serial 60 IS the phantom date, and
# from 61 on the extra day means the base becomes 1899-12-30.
_SERIAL_BASE_EARLY = date(1899, 12, 31)
_SERIAL_BASE_LATE  = date(1899, 12, 30)
# A bare integer in a date column is ambiguous: 2026 is both a plausible year and
# a valid serial (→ 1905-07-18). Rather than silently produce a 1905 date, only
# serials inside a sane window are accepted and anything else becomes a row
# ERROR quoting the raw value. 25569 = 1970-01-01, 73415 ≈ 2100-12-31.
_SERIAL_MIN, _SERIAL_MAX = 25569, 73415

# DAY-FIRST BEFORE MONTH-FIRST, and that ordering is load-bearing. Every date
# parser this codebase has ever had tried "%d/%m/%Y" before "%m/%d/%Y", so
# day-first is what the dates already in the database MEAN. "03/04/2026" reads as
# 3 April. Reversing the pair would not just change new imports, it would change
# how existing feeds are re-read, silently, with no error anywhere to notice.
#
# The list is deliberately long. The alternative — five separate parsers each
# accepting a different subset, which is what this codebase had — meant an
# invoice whose date arrived in an unlisted form was stored BLANK, and blank is
# indistinguishable from "the source sent nothing". Widening acceptance here can
# only turn an error or a blank into a date; a value that already parsed still
# parses the same way, because a format is only reached when every format before
# it has failed.
_DATE_FORMATS = (
    # Four-digit year, unambiguous separators.
    "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
    # Four-digit year, numeric day/month. Day-first first, see above.
    "%d/%m/%Y", "%m/%d/%Y",
    "%d-%m-%Y", "%m-%d-%Y",
    "%d.%m.%Y", "%m.%d.%Y",
    # Named month, any separator.
    "%d-%b-%Y", "%d-%B-%Y", "%b-%d-%Y", "%B-%d-%Y",
    "%d %b %Y", "%d %B %Y", "%b %d %Y", "%B %d %Y",
    "%B %d, %Y", "%b %d, %Y",
    # Two-digit year. Last, so a four-digit year can never be read as one.
    "%d/%m/%y", "%m/%d/%y",
    "%d-%m-%y", "%d-%b-%y", "%d-%B-%y",
    "%d.%m.%y", "%d %b %y", "%d %B %y",
    # Compact, as a CSV column typed to text emits it. Not reachable by a real
    # Excel serial: the widest serial in the window is five digits, and this
    # needs eight.
    "%Y%m%d",
)

# Collapses whitespace hugging a hyphen ("20 - Dec - 2025" -> "20-Dec-2025",
# "21-February -2026" -> "21-February-2026") so the dd-Mon-yyyy / dd-Month-yyyy
# formats above still match. See parse_import_date.
_HYPHEN_SPACING = re.compile(r"\s*-\s*")

# An Excel serial that arrived as TEXT rather than as a number, which is what a
# CSV export of a workbook produces, and what openpyxl gives for a cell whose
# column was typed to text. The fractional form ("45785.5104") is a serial
# carrying a time of day. Anchored at four or five digits because the plausible
# window (_SERIAL_MIN.._SERIAL_MAX) has no other width; excel_serial_to_date
# still enforces the window, this only decides what to hand it.
_SERIAL_TEXT = re.compile(r"^\d{4,5}(?:\.\d+)?$")

# A trailing time of day, with optional fractional seconds and optional timezone.
# Split off so ONE list of date formats serves both the dated and the timestamped
# spellings of the same column, rather than a second list with a time suffix
# bolted onto every entry. The colon is required, which is what stops this from
# eating the year out of "08 May 2026".
_TIME_TAIL = re.compile(
    r"[T\s](\d{1,2}):(\d{2})(?::(\d{2})(?:\.\d+)?)?\s*"
    r"(?:Z|[+-]\d{2}:?\d{2}|[A-Z]{2,5})?\s*$",
    re.IGNORECASE,
)

# A leading weekday, as a calendar export writes it: "Fri, 08 May 2026".
_WEEKDAY_PREFIX = re.compile(
    r"^(?:mon|tue|tues|wed|weds|thu|thur|thurs|fri|sat|sun)[a-z]*\.?,?\s+",
    re.IGNORECASE,
)

# An ordinal suffix on the day: "8th May 2026" -> "8 May 2026".
_ORDINAL_DAY = re.compile(r"\b(\d{1,2})(?:st|nd|rd|th)\b", re.IGNORECASE)

# The spellings a spreadsheet or a JSON export uses to mean "no value". Treated
# as blank rather than as an error, because they ARE the source saying nothing.
# "0" is here for the CSV exports that write a zero into an empty date column.
_BLANK_TEXT = frozenset(("nan", "nat", "none", "null", "n/a", "na", "-", "--", "0"))


def excel_serial_to_date(serial):
    """None when the serial is outside the plausible window or is the phantom."""
    n = int(serial)
    if n == 60 or n < _SERIAL_MIN or n > _SERIAL_MAX:
        return None
    base = _SERIAL_BASE_EARLY if n <= 59 else _SERIAL_BASE_LATE
    return base + timedelta(days=n)


def parse_import_date(raw):
    """
    Returns (date|None, error|None). A blank cell is (None, None) — these columns
    are nullable and an import must NOT apply any create-path default; a row that
    arrived undated stays undated.

    THE PARSERS, AND WHY THIS ONE IS THE AUTHORITY
    (matching note in frontend/src/lib/importParse.js)

      frontend/src/lib/importParse.js  reads .xlsx with SheetJS `raw: false`, so
                                       cells arrive as DISPLAYED TEXT and a
                                       serial never reaches the server that way
      THIS function                    accepts both — serials (bounded, phantom
                                       day 60 rejected) and strings — and ERRORS
                                       on anything it cannot read
      _parse_date in events/views.py
        and book_event/views.py        six string formats, no serial support, and
                                       returns None on failure, so a column of
                                       unreadable dates is indistinguishable from
                                       a column of blanks

    FIVE MORE USED TO EXIST AND NOW CALL THIS
    webhooks/services.py:parse_webhook_date (ten formats), the sheet sync
    (five), import_bookings_csv and import_bookings_json (ONE each, "%d-%b-%Y")
    and import_booking_excel (pandas, and month-first, so it disagreed with
    every other one about what "03/04/2026" meant). All five returned None on a
    format they did not know and reported nothing, so an unreadable date column
    imported as an empty one. They now delegate here and surface the returned
    error — in the webhook processing notes, in the command output, or in
    import_issues.md. See the note on _DATE_FORMATS about day-first ordering.

    Accepting both representations is what lets `load_zoho_export` read a workbook
    SERVER-side via openpyxl — where serials do arrive raw — without the browser
    having to normalise them first. Anything reading import files should call this.
    """
    if raw is None:
        return None, None
    # datetime before date: datetime is a date subclass.
    if isinstance(raw, datetime):
        return raw.date(), None
    if isinstance(raw, date):
        return raw, None
    # bool is an int subclass; reject before the numeric branch accepts it.
    if isinstance(raw, bool):
        return None, f"{raw!r} is not a date"
    if isinstance(raw, (int, float)):
        # NaN and +-inf are what pandas hands back for an empty or malformed
        # numeric cell, and int() RAISES on both. This branch used to reach
        # int(raw) unguarded, so a NaN in a date column crashed the parser
        # rather than being read as the blank it is.
        if isinstance(raw, float):
            if raw != raw:                      # NaN, which equals nothing
                return None, None
            if raw in (float("inf"), float("-inf")):
                return None, f"{raw!r} is not a recognisable date or Excel serial"
        # Excel fractional serials carry a time; the date part is what we want.
        resolved = excel_serial_to_date(int(raw))
        if resolved is None:
            return None, f"{raw!r} is not a recognisable date or Excel serial"
        return resolved, None

    # A non-breaking space is what a copy-paste out of a browser table leaves
    # behind, and it is not what str.strip() removes.
    text = " ".join(str(raw).replace(" ", " ").split())
    if not text:
        return None, None
    if text.lower() in _BLANK_TEXT:
        return None, None
    # A serial that arrived as text rather than as a number, including the
    # fractional form openpyxl produces for a cell carrying a time. Read here so
    # the format list below never sees it.
    if _SERIAL_TEXT.match(text):
        resolved = excel_serial_to_date(int(float(text)))
        if resolved is not None:
            return resolved, None
    # Strip a trailing time of day. This covers numpy/pandas datetime64, which
    # stringifies as "2026-08-10T00:00:00.000000000", and equally
    # "08/05/2026 10:30:00" and "2026-05-08T10:30:00Z", neither of which the
    # earlier ISO-only handling reached. Done BEFORE the hyphen collapse below,
    # so a negative timezone offset is gone before that rule can see its hyphen.
    text = _TIME_TAIL.sub("", text).strip()
    text = _WEEKDAY_PREFIX.sub("", text).strip()
    text = _ORDINAL_DAY.sub(r"\1", text)
    text = text.rstrip(",").strip()
    if not text:
        return None, f"{raw!r} is not a recognisable date"
    # Dirty spreadsheet exports from this same Zoho instance carry stray
    # whitespace around the hyphen in dd-Mon-yyyy style dates — "20 - Dec - 2025",
    # "21-February -2026" — which no entry in _DATE_FORMATS matches as typed.
    # Collapsed here so "%d-%b-%Y" / "%d-%B-%Y" still match. A no-op on every
    # other format this function accepts: ISO ("2026-08-10") and slash dates
    # ("10/08/2026") carry no whitespace around their own separators, so this can
    # only ever help the hyphenated formats, never change one that already parses.
    text = _HYPHEN_SPACING.sub("-", text)
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            continue
    return None, f"{raw!r} is not a recognisable date"


def parse_import_datetime(raw):
    """
    parse_import_date(), but keeping the time of day where the source sent one.

    Returns (datetime|None, error|None), same contract: blank is (None, None) and
    nothing raises. The datetime is NAIVE. Making it timezone-aware is the
    caller's job, because only the caller knows whether its source wrote local
    time or UTC, and guessing here would shift every backdated created_at by the
    UTC offset.

    Only the "Added Time" style columns need this — the ones that backdate
    BookEvent.created_at. Everything else wants the date and should call
    parse_import_date.
    """
    if raw is None:
        return None, None
    if isinstance(raw, datetime):
        return raw, None
    if isinstance(raw, bool):
        return None, f"{raw!r} is not a datetime"
    if isinstance(raw, date):
        return datetime(raw.year, raw.month, raw.day), None

    if isinstance(raw, (int, float)):
        # Same NaN/inf guard as parse_import_date; int() raises on both.
        if isinstance(raw, float):
            if raw != raw:
                return None, None
            if raw in (float("inf"), float("-inf")):
                return None, f"{raw!r} is not a recognisable datetime or Excel serial"
        resolved = excel_serial_to_date(int(raw))
        if resolved is None:
            return None, f"{raw!r} is not a recognisable datetime or Excel serial"
        # The fractional part of a serial IS the time of day, as a fraction of a
        # 24-hour day. parse_import_date discards it; here it is the point.
        seconds = round((float(raw) - int(raw)) * 86400)
        return datetime(resolved.year, resolved.month, resolved.day) + timedelta(
            seconds=seconds
        ), None

    text = " ".join(str(raw).replace(" ", " ").split())
    if not text:
        return None, None
    if text.lower() in _BLANK_TEXT:
        return None, None
    if _SERIAL_TEXT.match(text):
        return parse_import_datetime(float(text))

    # The date half goes through the one authority; only the time is read here,
    # so there is no second list of formats to drift out of step with the first.
    match = _TIME_TAIL.search(text)
    parsed_date, error = parse_import_date(text)
    if parsed_date is None:
        return None, error and error.replace("date", "datetime", 1)

    hour = minute = second = 0
    if match:
        hour, minute = int(match.group(1)), int(match.group(2))
        second = int(match.group(3) or 0)
        if not (hour <= 23 and minute <= 59 and second <= 59):
            # A 12-hour clock with an am/pm marker, or plain nonsense. Rather
            # than guess, keep the date and drop the time — a midnight timestamp
            # on the right day beats an error that loses the row's date too.
            hour = minute = second = 0

    return datetime(
        parsed_date.year, parsed_date.month, parsed_date.day, hour, minute, second
    ), None


# ── Editions ─────────────────────────────────────────────────────────────────
# An edition is a YEAR. The window is deliberately wide enough to be obviously
# safe (no real edition falls outside it) and narrow enough to catch the actual
# failure: an Excel serial. Serial 45678 is 2025-01-15 as a date, but read as an
# integer it is a plausible-looking 45678th edition, and `edition` is an
# IntegerField so nothing raises — it just stores and is silently wrong forever.
EDITION_MIN, EDITION_MAX = 2000, 2100


def parse_edition(raw):
    """
    Returns (int|None, error|None). Blank is (None, None) — edition is nullable
    and an absent edition is not an error.

    Two-digit years are expanded ("26" -> 2026) to match the convention
    BookEvent.save() already uses when it strips a year off an event code.
    """
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, None
    if isinstance(raw, bool):
        return None, f"{raw!r} is not an edition year"

    value, error = as_int(raw)
    if error:
        return None, f"{raw!r} is not an edition year"
    if value is None:
        return None, None

    if 0 <= value <= 99:
        value = 2000 + value

    if not (EDITION_MIN <= value <= EDITION_MAX):
        return None, (f"{raw!r} is not a plausible edition year "
                      f"(expected {EDITION_MIN}-{EDITION_MAX}) — a value this far "
                      f"out is usually an Excel serial date")
    return value, None


def clean_header(name):
    """Trimmed, whitespace-collapsed, lower-cased — the header lookup key."""
    return " ".join(str(name or "").strip().split()).lower()


def build_header_mapper(zoho_headers, model_fields):
    """
    Returns a map_headers(columns) -> (mapping, unrecognised) bound to one app's
    header table.

    Unrecognised columns are REPORTED, never silently dropped — a mistyped header
    would otherwise present as an entire column of empty data.
    """
    def map_headers(columns):
        mapping, unrecognised = {}, []
        for col in columns:
            key = clean_header(col)
            if key in zoho_headers:
                mapping[col] = zoho_headers[key]
            elif key.replace(" ", "_") in model_fields:
                mapping[col] = key.replace(" ", "_")
            else:
                unrecognised.append(col)
        return mapping, unrecognised
    return map_headers


def as_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def as_int(value):
    """Returns (int|None, error|None)."""
    if value is None:
        return None, None
    if isinstance(value, bool):
        return None, f"{value!r} is not a whole number"
    if isinstance(value, int):
        return value, None
    if isinstance(value, float):
        if not value.is_integer():
            return None, f"{value!r} is not a whole number"
        return int(value), None
    text = str(value).strip().replace(",", "")
    if not text:
        return None, None
    try:
        return int(text), None
    except ValueError:
        try:
            f = float(text)
        except ValueError:
            return None, f"{value!r} is not a whole number"
        if not f.is_integer():
            return None, f"{value!r} is not a whole number"
        return int(f), None


def as_bool(value):
    """
    Returns (bool, error|None). Spreadsheet booleans arrive as "Yes"/"No",
    "TRUE"/"FALSE", 1/0 or a real bool depending on which tool wrote the file.

    A blank cell is False rather than an error: these columns are
    BooleanField(default=False), so "absent" and "not set" are the same state and
    refusing the row would be inventing a requirement the form does not have.
    """
    if value is None or value == "":
        return False, None
    if isinstance(value, bool):
        return value, None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return bool(value), None
    text = str(value).strip().lower()
    if text in ("yes", "y", "true", "t", "1"):
        return True, None
    if text in ("no", "n", "false", "f", "0", "-"):
        return False, None
    return False, f"{value!r} is not a yes/no value"


# ── Cells that arrived as an anchor tag ──────────────────────────────────────
# Zoho writes some columns as HTML rather than as a bare value, so a LinkedIn
# cell reaches an import as
#     <a href="https://www.linkedin.com/in/eli-jasso" target="_blank">Eli</a>
# rather than as the address on its own. Stored verbatim that is not a link at
# all, on three counts. The grid renders href="&lt;a href=..." and the click goes
# nowhere; the markup eats the 500-character column, so a genuinely short URL can
# report as over-length; and a CSV export writes the tags back out, so the next
# import inherits them. The address inside the tag is the value the column was
# always meant to hold, so that is what gets stored, and everything downstream
# then treats the column exactly as it treats a hand-typed URL.
#
# ONLY http AND https SURVIVE. These cells come from an uploaded spreadsheet and
# the frontend renders the stored value straight into an href, so a
# `javascript:` address extracted here would sit one click from running in the
# CRM's own origin. frontend/src/lib/helpers.js:extUrl applies the same rule at
# render time. Both places keep it on purpose; the frontend cannot trust what is
# already in the database, and the database should not be storing what the
# frontend will refuse to render.
_ANCHOR_TAG = re.compile(r"<a\b", re.IGNORECASE)
_ANCHOR_HREF = re.compile(
    r"""<a\b[^>]*?\bhref\s*=\s*(?:"([^"]*)"|'([^']*)'|([^\s"'>]+))""",
    re.IGNORECASE,
)
_ANY_TAG = re.compile(r"<[^>]*>")
_HAS_SCHEME = re.compile(r"^[a-z][a-z0-9+.\-]*:", re.IGNORECASE)
_WEB_SCHEME = re.compile(r"^https?:", re.IGNORECASE)
_MAIL_OR_TEL = re.compile(r"^(?:mailto|tel):", re.IGNORECASE)
_HOSTLIKE = re.compile(r"^[\w-]+(?:\.[\w-]+)+$")


def unwrap_anchor(value):
    """
    (href, visible_text) for a cell written as anchor markup.

    `href` is None when the cell is not markup at all, which is the ordinary
    case; it is "" when there is an anchor carrying no usable href, such as
    `<a name="x">text</a>`. That three-way answer is what lets a caller tell
    "this was meant to be a link and the address is missing" apart from "this was
    never a link", and treat only the first as a row error.

    `visible_text` is the cell with every tag removed and HTML entities decoded,
    so `&amp;` in a tracking parameter comes back as `&`.
    """
    text = as_text(value)
    if not _ANCHOR_TAG.search(text):
        return None, text
    match = _ANCHOR_HREF.search(text)
    href = ""
    if match:
        # Exactly one group matches, whichever quoting style the tag used.
        href = unescape(next(g for g in match.groups() if g is not None)).strip()
    visible = " ".join(unescape(_ANY_TAG.sub(" ", text)).split())
    return href, visible


def absolute_url(value):
    """
    An http/https address a browser can actually navigate to, or None when the
    text is not a web address.

    A value with no scheme ("linkedin.com/in/x") is a RELATIVE path, so a browser
    resolves it against the CRM's own origin and the link reloads the CRM instead
    of going anywhere; it gets https:// only when what precedes the first slash
    plausibly names a host, so "N/A" and free-text notes come back None rather
    than becoming https://<prose>.

    Narrower than the frontend's extUrl, which also passes mailto: and tel: —
    those are legitimate to RENDER from a free-text column, but the two columns
    this guards are URLField, where Django's own URLValidator would reject them.
    """
    text = as_text(value)
    if not text or text == "—":
        return None
    if _HAS_SCHEME.match(text):
        return text if _WEB_SCHEME.match(text) else None
    if text.startswith("//"):
        return "https:" + text
    host = re.split(r"[/?#]", text, maxsplit=1)[0]
    return "https://" + text if _HOSTLIKE.match(host) else None


def as_url(value):
    """
    (url, error|None) for a column that stores a link.

    Anchor markup collapses to its address. Text that is already a URL passes
    through, gaining a scheme if it plausibly needs one.

    Text that is not a link at all is returned AS TYPED with no error, because
    "N/A" or "not on LinkedIn" in a LinkedIn column is a person answering the
    question, not a broken row, and failing the whole row over it would discard
    twenty good columns to police one. It renders as plain text rather than as a
    dead link — see ExtLink in frontend/src/components/UI.jsx. An anchor tag is
    the one case that DOES error: markup is unambiguously an attempt at a link,
    so an address that cannot be used is a real defect in the cell rather than a
    human writing prose.
    """
    raw = as_text(value)
    href, text = unwrap_anchor(raw)
    was_markup = href is not None

    candidate = (href or text) if was_markup else raw
    if not candidate:
        if was_markup:
            return "", "is an empty link tag, carrying no address at all"
        return "", None

    resolved = absolute_url(candidate)
    if resolved is not None:
        return resolved, None
    if was_markup:
        return "", (f"the address inside the link tag, {candidate!r}, is not an "
                    f"http or https address")
    return raw, None


def plain_text_cell(value):
    """
    A cell for a column that stores TEXT, with any anchor markup unwrapped.

    The address is kept alongside the visible text rather than thrown away, so a
    remark reading `<a href="https://x.com/deck">the deck</a>` imports as
    "the deck, https://x.com/deck" and the link is still there to follow. It is
    dropped only when it adds nothing: an address the visible text already
    contains, which is the shape Zoho writes most often, or a mailto:/tel: whose
    own address is the text, which is how an email column arrives.

    A cell with NO anchor in it comes back untouched, tags and all. That is a
    deliberate limit, not an oversight. Stripping every `<...>` from every text
    column would also eat the angle brackets people actually type — "<not stated>"
    in a remarks column is a value, not markup — and an anchor is the only shape
    with a demonstrated problem, because it is the only one whose contents the
    frontend is asked to navigate to. Widen this only against a real file that
    needs it.
    """
    href, text = unwrap_anchor(value)
    if href is None:
        return text
    bare = _MAIL_OR_TEL.sub("", href)
    if not text:
        return bare or href
    if bare and bare not in text:
        return f"{text}, {bare}"
    return text


def normalise_row(raw_row, mapping):
    """Map one inbound dict onto model field names."""
    out = {}
    for column, field in mapping.items():
        if column not in raw_row:
            continue
        out[field] = raw_row[column]
    return out


# ── Reading an actual file, server-side ──────────────────────────────────────
# The two-phase importers are fed JSON rows by the browser and never see a file.
# A management command that wants to run a REAL spreadsheet through the same
# classification needs to produce that same list-of-dicts shape itself, which is
# what this does.
#
# WHAT THIS SEES THAT THE BROWSER DOES NOT
# frontend/src/lib/importParse.js reads .xlsx with SheetJS `raw: false`, so every
# cell arrives as DISPLAYED TEXT and an Excel serial never reaches the server by
# that route. openpyxl with data_only=True returns the TYPED value: a real
# datetime where Excel stored a date, an int where it stored a serial. Both are
# handled by parse_import_date above, which is why routing every date through it
# is what makes a server-side read safe. The consequence worth stating plainly:
# reading a workbook here and uploading the same workbook through the browser are
# NOT identical operations — this path sees more of the truth.
#
# book_event/management/commands/load_zoho_export.py carries an equivalent private
# reader (_read_xlsx / _read_csv / _read_json). It is deliberately left alone
# rather than repointed here: it is covered by 32 passing tests and rewiring it is
# a refactor nobody asked for. This is the shared home for the next caller.
def read_import_rows(path):
    """
    Read .xlsx/.xlsm/.csv/.json into the list-of-dicts the importers expect.

    Keys are the file's own header labels, verbatim and untrimmed — header
    cleaning is map_headers' job, and trimming here would hide the fact that a
    label like "Closeness to Topic (10) " carries a trailing space.

    Rows that are entirely blank are skipped: a spreadsheet with formatting
    applied below the data yields hundreds of all-None tuples, and each one would
    otherwise become an ERROR row for three missing required fields.
    """
    lowered = str(path).lower()
    if lowered.endswith(".json"):
        with open(path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
        return data if isinstance(data, list) else []
    if lowered.endswith((".xlsx", ".xlsm")):
        return _read_xlsx_rows(path)
    return _read_csv_rows(path)


def _read_csv_rows(path):
    import csv
    # utf-8-sig: Excel writes a BOM, and without this the first header becomes
    # "﻿Event Code" and reports as an unrecognised column.
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _read_xlsx_rows(path):
    # data_only=True so a formula cell yields its cached VALUE rather than the
    # formula text. read_only=True to stream rather than materialise the sheet.
    from openpyxl import load_workbook

    book = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            header = [as_text(h) for h in next(rows)]
        except StopIteration:
            return []
        out = []
        for values in rows:
            if values is None or all(v is None or as_text(v) == "" for v in values):
                continue
            out.append({
                header[i]: values[i]
                for i in range(min(len(header), len(values)))
                if header[i] != ""
            })
        return out
    finally:
        book.close()


def plan_hash(plan):
    """
    Fingerprint of exactly what a commit would write: the normalised payload of
    every importable row, in row order, each including its RESOLVED event_code.

    Shape follows accounts/bulk_update.py:152 (sha256 over sorted, tight JSON) so
    the two behave identically for the caller, but computed here — that method's
    signature describes a single-field mass update and cannot express a row set.

    ERROR rows are excluded on purpose: they are never written, so a change to one
    must not invalidate an otherwise-current plan.
    """
    digest_input = [
        {"row": e["row"], "payload": e["_payload"]}
        for e in plan if e["classification"] != ERROR
    ]
    payload = json.dumps(digest_input, sort_keys=True, separators=(",", ":"),
                         default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def summarise(plan):
    counts = {CREATE: 0, CREATE_WITH_WARNING: 0, ERROR: 0}
    for entry in plan:
        counts[entry["classification"]] += 1
    return counts


def public_plan(plan):
    """Strip the internal payload before returning the plan to the client."""
    return [{k: v for k, v in entry.items() if k != "_payload"} for entry in plan]


# ── Does the value FIT its column ────────────────────────────────────────────
# Postgres integer column ranges, by Django field class name. Hard-coded rather
# than read off the field because Django exposes no such attribute, and these are
# fixed by the SQL types smallint, integer and bigint, not by anything settable.
_INT_BOUNDS = {
    "PositiveSmallIntegerField": (0, 32767),
    "SmallIntegerField":         (-32768, 32767),
    "PositiveIntegerField":      (0, 2147483647),
    "IntegerField":              (-2147483648, 2147483647),
    "PositiveBigIntegerField":   (0, 9223372036854775807),
    "BigIntegerField":           (-9223372036854775808, 9223372036854775807),
}


def column_errors(model, values, field_to_label):
    """
    Values that will not fit their database column, as ERROR entries.

    THE 500 THIS ENDS
    Both importers classify a row on business rules alone, meaning required fields,
    event code, dates and numeric ranges. Commit then writes the payload with a
    plain save(), which runs NO field validation; full_clean() is a serializer step
    and a form step, never a save step. So a value that is merely too WIDE passes
    preview as CREATE and only fails at the database, as a psycopg DataError inside
    the commit's transaction.atomic(). Two consequences follow, both bad.

      * It is a 500, not a row error, so the caller gets a debug page rather than a
        sentence naming the offending row and column.
      * The atomic block rolls the WHOLE chunk back, so 499 perfectly good rows are
        discarded because of one 255-character cell.

    It happened for real, on both importers at once. 'All Paper Reviews.csv' carries
    355 rows graded 'B+' against a 1-character column, and both exports carry two
    rows where an outreach message was pasted into Speaker Name, 255 characters
    against 150. Every commit of those files returned 500 and imported nothing.

    Checked against the MODEL rather than a hand-kept table of widths, so this
    cannot drift from the schema; widening a column relaxes this automatically.

    `values` should mirror the payload the commit will write, so pass the RESOLVED
    event code rather than the raw cell. Otherwise a long raw code that resolves to
    a short canonical one is reported as an error it will not actually cause.

    Sorted by label because both callers build `values` from a set, whose iteration
    order varies between processes; without this the same bad row reports its
    problems in a different order on each run.
    """
    from django.core.exceptions import FieldDoesNotExist

    errors = []
    for name, value in values.items():
        try:
            field = model._meta.get_field(name)
        except FieldDoesNotExist:
            continue

        label = field_to_label.get(name, name)

        if isinstance(value, str):
            limit = getattr(field, "max_length", None)
            if limit and len(value) > limit:
                errors.append({
                    "field": label,
                    "problem": f"longer than {limit} characters ({len(value)})",
                    "value": value[:80] + "…" if len(value) > 80 else value,
                })
        # bool is an int subclass and BooleanField has no range; exclude it.
        elif isinstance(value, int) and not isinstance(value, bool):
            bounds = _INT_BOUNDS.get(type(field).__name__)
            if bounds and not (bounds[0] <= value <= bounds[1]):
                errors.append({
                    "field": label,
                    "problem": f"outside the range this column stores "
                               f"({bounds[0]} to {bounds[1]})",
                    "value": str(value),
                })

    return sorted(errors, key=lambda e: e["field"])


def catalogue_notice():
    """
    A whole-file explanation when the Events catalogue is EMPTY, else None.

    THE CONFUSION THIS ENDS
    Both importers resolve every row's Event Code against `events`, so with an empty
    catalogue every single row comes back ERROR "no matching event; prefilter
    candidates []". That is accurate per row and useless as a diagnosis: a 400-row
    file returns 400 identical errors, the Import button stays disabled because
    nothing is importable, and nothing anywhere says the catalogue is the problem.
    It happened for real — the catalogue was cleared and the next import read as a
    broken import button.

    Returned as a top-level notice rather than folded into the per-row errors,
    because it is a fact about the SYSTEM, not about any row in the file.
    """
    from events.models import Event

    if Event.objects.exists():
        return None
    return (
        "The Events catalogue is empty, so no row's Event Code can be matched and "
        "nothing in this file can be imported. Restore the events first (Events → "
        "Import), then import this file again."
    )
