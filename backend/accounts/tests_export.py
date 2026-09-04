"""
accounts/tests_export.py
─────────────────────────
AdminExportMixin — `GET {resource}/export/`.

THREE PROPERTIES, AND ONLY THREE ARE WORTH PINNING

  * THE FILE HOLDS WHAT THE TABLE HOLDS, no more. The first cut of this export
    wrote every field the serializer reports, which on Bookings meant `id`,
    `book_event_id`, `accounts_contact_email_raw` and fourteen delegate_* /
    effective_* columns nobody has ever seen on screen. The header assertion
    below is exact, and being exact is the point, a subset check would pass
    again on the day somebody restores the "just use the serializer" default.

  * THE FILE HOLDS THE FILTERED ROWS, NOT THE TABLE. An export answering a
    wider query than the screen it was taken from is a data leak with a
    filename attached, and it is invisible from the UI, the file downloads, it
    opens, it has rows in it. So these compare against the LIST endpoint's own
    answer to the same query rather than a hand-counted number.

  * ADMIN ONLY, ON TOP OF THE MODULE GATE. Bookings view is not the right to
    walk out with every booking in one workbook, and the gate has to hold for a
    caller holding every Bookings cell there is.

    python manage.py test accounts.tests_export
"""
import io
import json
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from urllib.parse import quote

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIRequestFactory, force_authenticate

from book_delegate.models import BookDelegate
from book_delegate.views import BookDelegateViewSet
from book_event.models import BookEvent
from teams.models import Team, TeamPermission

User = get_user_model()

EXPORT = BookDelegateViewSet.as_view({"get": "export"})
LIST = BookDelegateViewSet.as_view({"get": "list"})

KEPT = "EXP - AA"
DROPPED = "EXP - ZZ"

# frontend/src/pages/BookingsPage.jsx bkCols, in order, minus the Transfer
# button. Written out here rather than read off the viewset, or the test would
# agree with the code by construction and assert nothing.
TABLE_HEADERS = [
    "Payment Status", "Event Code", "Booking Code", "Request Date",
    "Invoice Date", "Invoice Number", "Name", "Delegate Company",
    "Delegate Email", "Direct Line", "Accounts Contact", "Delegate Number",
    "Payable/Free", "Date Paid", "Payment Type", "Ticket Tier", "Discount",
    "Add-Ons", "Ref", "Event Name", "Added Time", "Modified Time",
    "Sales Executive", "Attendance - IN?",
]


class AdminExportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.all_access = Team.objects.create(name="exp_admin", is_all_access=True)
        cls.admin = User.objects.create_user(
            username="exp_admin", password="x", email="exp_admin@iq-hub.com",
            role="admin", team=cls.all_access)

        # role="sales", NOT on an all-access team, holding every cell the
        # Bookings grid can give. The whole point, full module access is not
        # admin, and this caller must still be refused.
        cls.granted_team = Team.objects.create(name="exp_sales")
        cls.granted = User.objects.create_user(
            username="exp_sales", password="x", email="exp_sales@iq-hub.com",
            role="sales", team=cls.granted_team)
        TeamPermission.objects.update_or_create(
            team=cls.granted_team, module="bookings",
            defaults=dict(can_view=True, can_create=True,
                          can_update=True, can_delete=True))

        today = date.today()
        cls.kept_invoice = BookEvent.objects.create(
            invoice_number="EXP-KEPT", event_code=KEPT,
            payment_status="Paid", request_date=today)
        cls.dropped_invoice = BookEvent.objects.create(
            invoice_number="EXP-DROP", event_code=DROPPED,
            payment_status="Paid", request_date=today - timedelta(days=400))

        for i in range(3):
            BookDelegate.objects.create(
                invoice=cls.kept_invoice, event_code=KEPT,
                first_name="Kept", last_name=str(i),
                email="kept{}@example.com".format(i),
                discount=Decimal("0.20"))
        BookDelegate.objects.create(
            invoice=cls.dropped_invoice, event_code=DROPPED,
            first_name="Dropped", last_name="0", email="dropped@example.com")

    def setUp(self):
        self.factory = APIRequestFactory()

    # ── helpers ──────────────────────────────────────────────────────────────

    def _get(self, view, query="", user=None):
        req = self.factory.get("/?" + query)
        force_authenticate(req, user=user or self.admin)
        return view(req)

    def _sheet(self, query="", user=None):
        """The workbook as (header, list-of-dicts)."""
        from openpyxl import load_workbook

        resp = self._get(EXPORT, query, user)
        self.assertEqual(resp.status_code, 200, getattr(resp, "data", b""))
        book = load_workbook(io.BytesIO(resp.content), read_only=True)
        rows = list(book.active.iter_rows(values_only=True))
        header = list(rows[0])
        return header, [dict(zip(header, r)) for r in rows[1:]]

    def _listed_emails(self, query=""):
        tail = query + "&page_size=500" if query else "page_size=500"
        resp = self._get(LIST, tail)
        resp.render()
        self.assertEqual(resp.status_code, 200, resp.content)
        return sorted(r["email"] for r in json.loads(resp.content)["results"])

    @staticmethod
    def _spec(criteria):
        return "filter_spec=" + quote(json.dumps({"match": "all", "criteria": criteria}))

    # ── The columns ──────────────────────────────────────────────────────────

    def test_it_returns_a_workbook_named_for_the_module(self):
        resp = self._get(EXPORT)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn('filename="bookings.xlsx"', resp["Content-Disposition"])
        # A real .xlsx is a zip, and PK is its magic number. Not a renamed CSV.
        self.assertEqual(resp.content[:2], b"PK")

    def test_the_header_is_the_tables_columns_and_nothing_else(self):
        """THE REGRESSION. The export used to carry the whole API payload."""
        header, _ = self._sheet()
        self.assertEqual(header, TABLE_HEADERS)

    def test_no_internal_field_reaches_the_file(self):
        """
        Named individually because these are the ones that actually leaked, and
        the header assertion above would not say WHICH if it broke.
        """
        header, rows = self._sheet()
        flat = " ".join(header).lower()
        for internal in ("book event", "raw", "effective", "delegate payment",
                         "created at", "updated at", "paid free"):
            self.assertNotIn(internal, flat)
        # And no stray keys smuggled in on the rows themselves.
        self.assertEqual(set(rows[0]), set(TABLE_HEADERS))

    def test_discount_is_the_percent_the_table_shows(self):
        """0.20 stored, 20 on screen, so 20 in the cell, not 0.2."""
        _, rows = self._sheet(
            self._spec([{"field": "event_code", "op": "is", "value": KEPT}]))
        self.assertEqual({r["Discount"] for r in rows}, {20.0})

    def test_timestamps_are_real_date_cells_read_in_ist(self):
        """
        Not "2026-08-25T20:26:32.336950Z" in a text cell. The table renders
        every timestamp at +05:30 (frontend/src/lib/helpers.js), and a workbook
        holding the UTC instant would date anything logged after 18:30 to the
        previous day, against a screen saying otherwise.
        """
        _, rows = self._sheet(
            self._spec([{"field": "event_code", "op": "is", "value": KEPT}]))
        cell = rows[0]["Added Time"]
        self.assertIsInstance(cell, datetime)
        self.assertIsNone(cell.tzinfo, "Excel has no offsets; openpyxl refuses aware")

        row = BookDelegate.objects.get(email=rows[0]["Delegate Email"])
        ist = row.created_at.astimezone(timezone(timedelta(hours=5, minutes=30)))
        self.assertEqual(cell, ist.replace(tzinfo=None, microsecond=0))

    def test_a_plain_date_is_not_shifted(self):
        """
        A DateField is a calendar day with no instant behind it. Moving it five
        and a half hours would invent a timezone for a value that has none, and
        would move every date booked before 05:30 to the day before.
        """
        _, rows = self._sheet(
            self._spec([{"field": "event_code", "op": "is", "value": KEPT}]))
        cell = rows[0]["Request Date"]
        self.assertEqual(
            cell.date() if isinstance(cell, datetime) else cell, date.today())

    # ── The rows ─────────────────────────────────────────────────────────────

    def test_it_exports_the_filtered_rows_and_only_those(self):
        """Same query in, same rows out as the list."""
        query = self._spec([{"field": "event_code", "op": "is", "value": KEPT}])

        _, rows = self._sheet(query)

        self.assertEqual(sorted(r["Delegate Email"] for r in rows),
                         self._listed_emails(query))
        self.assertEqual(len(rows), 3)
        self.assertNotIn("dropped@example.com", [r["Delegate Email"] for r in rows])

    def test_the_period_window_narrows_it_too(self):
        """
        `?period=` is not a filter_spec criterion, it is PeriodFilterMixin, and
        it applies only to the actions named in `period_actions`. Left out,
        exporting from inside "Last 30 days" would quietly be an export of
        everything, the file opens, it has rows in it, and it is wrong.
        """
        _, rows = self._sheet("period=last_30_days")

        self.assertEqual(sorted(r["Delegate Email"] for r in rows),
                         self._listed_emails("period=last_30_days"))
        self.assertNotIn("dropped@example.com", [r["Delegate Email"] for r in rows])

    # ── The gate ─────────────────────────────────────────────────────────────

    def test_every_bookings_grant_is_still_not_admin(self):
        resp = self._get(EXPORT, user=self.granted)
        self.assertEqual(resp.status_code, 403)

    def test_the_module_gate_survives_the_admin_gate(self):
        """
        IsAdminRole is APPENDED, not swapped in. A caller with no Bookings
        access at all is refused here as everywhere else, the two gates are
        AND, and a `permission_classes` on the action would have made the
        second one replace the first.
        """
        outsider = User.objects.create_user(
            username="exp_outsider", password="x", email="exp_out@iq-hub.com",
            role="sales", team=Team.objects.create(name="exp_none"))

        self.assertEqual(self._get(EXPORT, user=outsider).status_code, 403)
        self.assertEqual(self._get(LIST, user=outsider).status_code, 403)


class DeclaredColumnsTests(TestCase):
    """
    Every viewset mixing the export in must declare columns its serializer can
    actually produce. A typo here is a 500 for an admin at the moment they press
    Export, which is the worst possible time to find out.

    Discovered off AdminExportMixin's subclasses rather than a hardcoded list,
    so a module that takes the export on later is covered without editing this.
    """

    def test_every_declared_column_resolves(self):
        from accounts.spreadsheet_export import AdminExportMixin

        viewsets = AdminExportMixin.__subclasses__()
        self.assertIn(BookDelegateViewSet, viewsets)

        for viewset in viewsets:
            with self.subTest(viewset.__name__):
                self.assertTrue(viewset.export_columns,
                                "export_columns is required; there is no "
                                "'export everything' default")
                view = viewset()
                view.action = "export"
                view.format_kwarg = None
                serializer = view.get_serializer_class()()
                unknown = [f for f, _ in viewset.export_columns
                           if f not in serializer.fields
                           and f not in viewset.export_values]
                self.assertEqual(unknown, [])
