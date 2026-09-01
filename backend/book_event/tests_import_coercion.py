"""
book_event/tests_import_coercion.py
────────────────────────────────────
The eight fixes from the 26 August data-integrity review, one class each, each
asserting that review's own acceptance test.

    1 HeaderAliasTests           all 21 headers auto-map, unmappable is reported
    2 PayableSpellingTests       a row reading Payable imports as Paid
    3 PersonLevelWriteTests      one Free and one Payable delegate store both
    4 RejectUnreadableTests      an unreadable value errors its row and writes nothing
    5 PercentDiscountTests       20% and 0.2 store the same discount
    6 BatchIdentifierTests       an import's rows list from its batch id alone
    7 DryRunTests                per-column counts, and the preview writes nothing
    8 SharedCoercionTests        every write path coerces a value identically

Each class opens with the OLD behaviour written down, so the test records what was
wrong rather than only what is now right.

WHY THESE NUMBERS APPEAR IN THE ASSERTIONS
They are the measured blast radius from the review of
"Master Data to Fancy Google Sheet 26 Aug.xlsx", 15,180 rows over 11,288
invoices: 11,205 rows lost Payable/Free to an unrecognised spelling, 903 invoices
carried a mix of Free and Payable that was flattened to one value, 868 carried
more than one Booking Code, 671 discounts written "20%" became zero, and 4,636
zeros in Delegate Count were rewritten as ones. The fixtures below are small, but
they are shaped like those rows.

    python manage.py test book_event.tests_import_coercion
"""
from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIRequestFactory, force_authenticate

from accounts.booking_coercion import (
    RULES, UNSET, allowed_values, coerce, coerce_row, column_report,
    percent_to_fraction,
)
from book_delegate.models import BookDelegate
from book_event.models import BookEvent
from book_event.views import BOOKING_IMPORT_FIELDS, BookEventViewSet
from events.models import Event
from teams.models import Team

User = get_user_model()


# ── shared fixtures ─────────────────────────────────────────────────────────

def all_access_admin(username):
    """
    An admin who passes crm_permission(). role="admin" alone is not enough:
    accounts/crm_permissions.py reads user.team.is_all_access, and a user with
    no team is refused whatever their role says.
    """
    team, _ = Team.objects.get_or_create(
        name="import_coercion_admin", defaults={"is_all_access": True})
    user = User.objects.create_user(
        username=username, password="x", role=User.Role.ADMIN,
        email=f"{username}@iq-hub.com",
    )
    user.team = team
    user.save()
    return user


def post_import(user, rows, **body):
    view = BookEventViewSet.as_view({"post": "bulk_import"})
    request = APIRequestFactory().post("/", {"rows": rows, **body}, format="json")
    force_authenticate(request, user=user)
    response = view(request)
    assert response.status_code == 200, f"{response.status_code}: {response.data}"
    return response.data


def row(**over):
    """One delegate row, in the shape the wizard sends after mapping."""
    base = {
        "invoice_number": "INV-1",
        "event_code":     "TEST-26",
        "event_name":     "Test Event",
        "contact_name":   "Ada Lovelace",
        "contact_email":  "ada@example.test",
        "company_name":   "Analytical Engines",
    }
    base.update(over)
    return base


class ImportCase(TestCase):
    """Base with an admin and an event the resolver can find."""

    @classmethod
    def setUpTestData(cls):
        cls.admin = all_access_admin(f"imp.{cls.__name__[:12].lower()}")
        Event.objects.get_or_create(
            event_code="TEST-26",
            defaults={"official_event_name": "Test Event",
                      "event_date": date(2026, 8, 26)},
        )


# ══ 1 HEADER ALIASES ════════════════════════════════════════════════════════

class HeaderAliasTests(TestCase):
    """
    "Delegate Company" and "Delegate Email" resolved to NOTHING and were skipped,
    and a skipped column is indistinguishable in the wizard from a column the
    file never contained. Delegate Email is the delegate identity key, so losing
    it also meant every second row on an invoice was given a
    dup-xxxxxxxx@import.local placeholder address.
    """

    # The 21 headers of the reviewed file, VERBATIM, read off row 1 of
    # backend/data_imports/master_data_26aug.xlsx rather than reconstructed from
    # the review's prose. Two of them ("Name", "Accounts Contact") are not what a
    # reconstruction would have guessed, and a test that guesses its own input
    # proves nothing about the file that actually broke.
    FILE_HEADERS = (
        "Payment Status", "Event Code", "Booking Code", "Request Date",
        "Invoice Date", "Invoice Number", "Name", "Delegate Company",
        "Delegate Email", "Direct Line", "Accounts Contact", "Delegate Count",
        "Payable/Free", "Payment Date", "Payment Type", "Ticket Tier",
        "Discount", "Add-Ons", "Ref", "Event Name", "Attendance - IN?",
    )

    # Headers carried by the OTHER sheets we import, which the review also named.
    # Kept separate so the count above stays the reviewed file's own 21.
    OTHER_SHEET_HEADERS = ("Date Paid", "Currency", "Notes", "Added Time")

    @staticmethod
    def nrm(s):
        return "".join(c for c in str(s).lower() if c.isalnum())

    @classmethod
    def auto_map(cls, header):
        """
        A faithful port of autoMap in frontend/src/components/ImportWizard.jsx:
        exact match on key, label or alias across ALL fields first, then a
        symmetric substring scan, then skip.
        """
        norm = cls.nrm(header)
        for key, label, aliases in BOOKING_IMPORT_FIELDS:
            if norm in {cls.nrm(key), cls.nrm(label)} | {cls.nrm(a) for a in aliases}:
                return key
        for key, _, _ in BOOKING_IMPORT_FIELDS:
            kn = cls.nrm(key)
            if kn in norm or norm in kn:
                return key
        return None

    def test_all_21_headers_of_the_file_now_map(self):
        """The review's acceptance test for fix 1, against the real header row."""
        self.assertEqual(len(self.FILE_HEADERS), 21)
        unmapped = [h for h in self.FILE_HEADERS if self.auto_map(h) is None]
        self.assertEqual(unmapped, [], f"still unmappable: {unmapped}")

    def test_the_headers_match_the_workbook_still_on_disk(self):
        """
        Guards the fixture above against the file being replaced. Skips rather
        than fails where the workbook is not in the checkout, because it is data
        and not every clone carries it.
        """
        from pathlib import Path

        import openpyxl
        from django.conf import settings

        path = Path(settings.BASE_DIR) / "data_imports" / "master_data_26aug.xlsx"
        if not path.exists():
            self.skipTest("data_imports/master_data_26aug.xlsx not present")
        ws = openpyxl.load_workbook(path, read_only=True).worksheets[0]
        actual = tuple(
            str(h).strip() for h in next(ws.iter_rows(max_row=1, values_only=True))
            if h is not None
        )
        self.assertEqual(actual, self.FILE_HEADERS)

    def test_the_headers_the_other_sheets_carry_also_map(self):
        unmapped = [h for h in self.OTHER_SHEET_HEADERS if self.auto_map(h) is None]
        self.assertEqual(unmapped, [], f"still unmappable: {unmapped}")

    def test_the_two_columns_that_were_lost_map_to_the_right_fields(self):
        self.assertEqual(self.auto_map("Delegate Company"), "company_name")
        self.assertEqual(self.auto_map("Delegate Email"), "contact_email")
        self.assertEqual(self.auto_map("Date Paid"), "payment_date")
        self.assertEqual(self.auto_map("Ref"), "reference")
        self.assertEqual(self.auto_map("Delegate Count"), "delegate_count")
        self.assertEqual(self.auto_map("Attendance - IN?"), "attendance")

    def test_the_old_substring_scan_would_have_missed_them(self):
        """
        Records the defect. Without the alias, "delegatecompany" is compared
        against "companyname" and neither contains the other, so the loose scan
        found nothing and the column was skipped in silence.
        """
        self.assertNotIn("companyname", "delegatecompany")
        self.assertNotIn("delegatecompany", "companyname")
        self.assertNotIn("contactemail", "delegateemail")

    def test_a_deliberately_unmappable_column_is_still_unmappable(self):
        """
        The aliases must not be so loose that everything matches something. A
        column reported as unmapped is the wizard's remaining safety net, and it
        only works if the report can still happen.
        """
        self.assertIsNone(self.auto_map("Sponsor Lanyard Colour"))

    def test_the_accounts_email_column_is_not_stolen_by_the_new_alias(self):
        """
        "Delegate Email" is an alias on contact_email, and accounts_contact_email
        is declared BEFORE it for the substring scan's sake. Exact matching has to
        keep both columns pointing where they belong.
        """
        self.assertEqual(self.auto_map("Accounts Email"), "accounts_contact_email")
        self.assertEqual(self.auto_map("Email"), "contact_email")


# ══ 2 PAYABLE ═══════════════════════════════════════════════════════════════

class PayableSpellingTests(ImportCase):
    """
    Payable/Free was coerced through a map built from the model's two stored
    values, so only "paid" and "free" were recognised. "Payable" is the word the
    CRM DISPLAYS for "Paid" and no importer had ever accepted it, which is one
    line of code and 11,205 lost rows.
    """

    def test_a_row_reading_payable_imports_as_paid(self):
        """
        The review's acceptance test for fix 2. Checked as the RESOLVED value,
        `delegate_paid_or_free or invoice.paid_or_free`, because on a single-
        delegate invoice the value settles on the invoice and the override is
        cleared — see PersonLevelWriteTests. Asserting the override alone would
        pass or fail for reasons that have nothing to do with Payable.
        """
        post_import(self.admin, [row(paid_or_free="Payable")])
        delegate = BookDelegate.objects.select_related("invoice").get()
        resolved = delegate.delegate_paid_or_free or delegate.invoice.paid_or_free
        self.assertEqual(resolved, "Paid")
        self.assertEqual(BookEvent.objects.get().paid_or_free, "Paid")

    def test_the_display_label_and_the_accepted_input_agree(self):
        """
        The audit the review asked for, as an assertion: every value we relabel
        for display must be accepted as an input spelling. Read from the frontend
        constant so a future relabelling that is not carried back here fails.
        """
        from pathlib import Path

        from django.conf import settings

        js = Path(settings.BASE_DIR).parent / "frontend" / "src" / "lib" / "constants.js"
        if not js.exists():
            self.skipTest("frontend/src/lib/constants.js not present")
        import re
        src = js.read_text(encoding="utf-8")
        m = re.search(r"PAID_OR_FREE_LABEL\s*=\s*\{([^}]*)\}", src)
        self.assertIsNotNone(m, "PAID_OR_FREE_LABEL not found")
        for stored, shown in re.findall(r"(\w+)\s*:\s*'([^']+)'", m.group(1)):
            value, error = coerce("paid_or_free", shown)
            self.assertIsNone(
                error,
                f"the UI shows {shown!r} for {stored!r} but the importer rejects it",
            )
            self.assertEqual(value, stored)

    def test_the_other_spellings_the_repair_command_accepts(self):
        for spelling, stored in (
            ("payable", "Paid"), ("PAYABLE", "Paid"), ("Chargeable", "Paid"),
            ("Free", "Free"), ("free", "Free"), ("FOC", "Free"),
            ("Complimentary", "Free"), ("comp", "Free"),
        ):
            with self.subTest(spelling=spelling):
                self.assertEqual(coerce("paid_or_free", spelling), (stored, None))

    def test_a_blank_stays_blank_and_is_not_read_as_charged(self):
        self.assertEqual(coerce("paid_or_free", ""), ("", None))
        self.assertEqual(coerce("paid_or_free", None), ("", None))

    def test_an_unknown_spelling_is_reported_not_defaulted(self):
        value, error = coerce("paid_or_free", "Sponsored")
        self.assertIsNone(value)
        self.assertIn("Sponsored", error)
        self.assertIn("Free", error)
        self.assertIn("Paid", error)


# ══ 3 PERSON-LEVEL WRITES ═══════════════════════════════════════════════════

class PersonLevelWriteTests(ImportCase):
    """
    The file is one row per delegate. Payable/Free, Payment Status, Payment Type,
    Payment Date, Ticket Tier and Booking Code were written on the INVOICE only,
    so where delegates on one invoice differed, one row's value was applied to
    all of them and which row won depended on row order. 903 invoices in the
    reviewed file carry a mix of Free and Payable; 868 carry more than one
    Booking Code, which drives revenue classification.
    """

    MIXED = [
        row(contact_name="Ada Lovelace", contact_email="ada@example.test",
            paid_or_free="Free", booking_code="Speaker", payment_type="Bank",
            ticket_tier="SEB", payment_status="Free"),
        row(contact_name="Alan Turing", contact_email="alan@example.test",
            paid_or_free="Payable", booking_code="Delegate", payment_type="Stripe",
            ticket_tier="EB", payment_status="Paid"),
    ]

    def test_one_free_and_one_payable_delegate_both_store_correctly(self):
        """The review's acceptance test for fix 3, word for word."""
        post_import(self.admin, self.MIXED)

        ada = BookDelegate.objects.get(email="ada@example.test")
        alan = BookDelegate.objects.get(email="alan@example.test")
        self.assertEqual(ada.delegate_paid_or_free, "Free")
        self.assertEqual(alan.delegate_paid_or_free, "Paid")

    def test_the_bookings_table_shows_both(self):
        """
        What the table actually renders is the RESOLVED value,
        `delegate_paid_or_free or invoice.paid_or_free` — see
        book_delegate/serializers.py. Asserting the stored override is not enough:
        the defect was visible precisely because the resolved values agreed when
        they should have differed.
        """
        post_import(self.admin, self.MIXED)
        from book_delegate.serializers import BookDelegateInlineSerializer

        resolved = {
            d["email"]: d["effective_paid_or_free"]
            for d in BookDelegateInlineSerializer(
                BookDelegate.objects.select_related("invoice"), many=True).data
        }
        self.assertEqual(resolved["ada@example.test"], "Free")
        self.assertEqual(resolved["alan@example.test"], "Paid")

    def test_every_person_level_column_carries_its_own_value(self):
        post_import(self.admin, self.MIXED)
        ada = BookDelegate.objects.get(email="ada@example.test")
        alan = BookDelegate.objects.get(email="alan@example.test")
        self.assertEqual((ada.booking_code, alan.booking_code), ("Speaker", "Delegate"))
        self.assertEqual((ada.delegate_payment_type, alan.delegate_payment_type),
                         ("Bank", "Stripe"))
        self.assertEqual((ada.delegate_ticket_tier, alan.delegate_ticket_tier),
                         ("SEB", "EB"))
        self.assertEqual((ada.delegate_payment_status, alan.delegate_payment_status),
                         ("Free", "Paid"))

    def test_row_order_no_longer_decides_the_outcome(self):
        """
        The root cause of the randomness. The same two rows in the other order
        must produce the same stored data.
        """
        post_import(self.admin, self.MIXED)
        forwards = {d.email: d.delegate_paid_or_free for d in BookDelegate.objects.all()}

        BookDelegate.objects.all().delete()
        BookEvent.objects.all().delete()
        post_import(self.admin, list(reversed(self.MIXED)))
        backwards = {d.email: d.delegate_paid_or_free for d in BookDelegate.objects.all()}

        self.assertEqual(forwards, backwards)

    def test_the_invoice_takes_the_value_when_every_delegate_agrees(self):
        """
        The other half of the rule the booking modal already applies: an override
        exists to carry a DIFFERENCE, so where there is none the value belongs on
        the invoice and the overrides are cleared. Otherwise every
        invoice-level report drifts away from what the Bookings tab shows.
        """
        post_import(self.admin, [
            row(contact_email="a@example.test", contact_name="A One", paid_or_free="Free"),
            row(contact_email="b@example.test", contact_name="B Two", paid_or_free="Free"),
        ])
        self.assertEqual(BookEvent.objects.get().paid_or_free, "Free")
        self.assertEqual(
            list(BookDelegate.objects.values_list("delegate_paid_or_free", flat=True)),
            [None, None],
        )

    def test_a_mixed_invoice_leaves_the_invoice_column_alone(self):
        post_import(self.admin, self.MIXED)
        invoice = BookEvent.objects.get()
        # Nothing agreed, so nothing was pushed up; the overrides carry it.
        self.assertEqual(invoice.paid_or_free, "")

    def test_re_importing_the_same_rows_changes_nothing(self):
        """
        The review's verification step: re-importing the same file must change
        nothing. That is the test that the repair and the fix are both real.
        """
        post_import(self.admin, self.MIXED)
        before = sorted(BookDelegate.objects.values_list(
            "email", "delegate_paid_or_free", "booking_code",
            "delegate_payment_type", "delegate_ticket_tier"))
        invoice_before = BookEvent.objects.values().get()

        post_import(self.admin, self.MIXED, duplicate_strategy="upsert")

        after = sorted(BookDelegate.objects.values_list(
            "email", "delegate_paid_or_free", "booking_code",
            "delegate_payment_type", "delegate_ticket_tier"))
        self.assertEqual(before, after)
        self.assertEqual(BookDelegate.objects.count(), 2)
        invoice_after = BookEvent.objects.values().get()
        for key, value in invoice_before.items():
            # updated_at moves by definition; import_batch_id names the run that
            # last touched the row, which is the point of it.
            if key in ("updated_at", "import_batch_id"):
                continue
            self.assertEqual(invoice_after[key], value, f"{key} moved on re-import")

    def test_the_invoice_contact_follows_the_first_delegate_not_the_last_row(self):
        """
        The upsert path assigned contact_name from every row in turn, so on a
        four-delegate invoice the invoice's own Delegate Name was whichever row
        the file happened to list last — and it CHANGED on a re-import. It is now
        the first delegate's, which is what the website intake has always stored.
        """
        post_import(self.admin, self.MIXED)
        self.assertEqual(BookEvent.objects.get().contact_name, "Ada Lovelace")
        post_import(self.admin, list(reversed(self.MIXED)), duplicate_strategy="upsert")
        self.assertEqual(BookEvent.objects.get().contact_name, "Ada Lovelace")

    def test_the_invoice_delegate_count_is_derived_from_the_rows(self):
        post_import(self.admin, self.MIXED)
        self.assertEqual(BookEvent.objects.get().delegate_count, 2)


# ══ 4 REJECT UNREADABLE VALUES ══════════════════════════════════════════════

class RejectUnreadableTests(ImportCase):
    """
    A value the importer did not recognise was replaced with a blank, with a
    default, or left as whatever was stored — and nothing was reported. Every
    default the review named is checked here.
    """

    def test_an_unreadable_value_errors_its_row_and_writes_nothing(self):
        """The review's acceptance test for fix 4."""
        data = post_import(self.admin, [
            row(invoice_number="INV-GOOD", paid_or_free="Free"),
            row(invoice_number="INV-BAD", contact_email="bad@example.test",
                paid_or_free="Sponsored"),
        ])
        self.assertEqual(data["inserted"], 1)
        self.assertEqual(len(data["errors"]), 1)

        error = data["errors"][0]
        self.assertEqual(error["row_index"], 1)
        self.assertEqual(error["invoice_number"], "INV-BAD")
        self.assertIn("Payable / Free", error["message"])
        self.assertIn("Sponsored", error["message"])

        # Nothing partial: the bad row's invoice and delegate are both absent.
        self.assertFalse(BookEvent.objects.filter(invoice_number="INV-BAD").exists())
        self.assertFalse(BookDelegate.objects.filter(email="bad@example.test").exists())
        # And the good row is unaffected.
        self.assertEqual(BookEvent.objects.get().invoice_number, "INV-GOOD")

    def test_a_row_reports_every_bad_cell_at_once(self):
        """
        Fixing a spreadsheet one error per run is not a workflow anybody
        completes, so a row names all of its problems.
        """
        data = post_import(self.admin, [
            row(paid_or_free="Sponsored", ticket_tier="Platinum", currency="Dollars"),
        ])
        message = data["errors"][0]["message"]
        for expected in ("Payable / Free", "Ticket Tier", "Currency"):
            self.assertIn(expected, message)

    def test_a_genuinely_blank_cell_stays_blank(self):
        data = post_import(self.admin, [row(paid_or_free="", ticket_tier="")])
        self.assertEqual(data["errors"], [])
        self.assertEqual(BookEvent.objects.get().paid_or_free, "")

    def test_attendance_is_translated_not_defaulted(self):
        """
        "false" matched no recognised spelling, so 13,481 rows fell through to the
        Pending default. Pending is the right answer for an unticked flag, but it
        has to be REACHED rather than fallen into — and the same fallback silently
        absorbed Absent, which means something else entirely.
        """
        self.assertEqual(coerce("attendance", "true"), ("Confirmed", None))
        self.assertEqual(coerce("attendance", "false"), ("Pending", None))
        self.assertEqual(coerce("attendance", "Absent"), ("No-show", None))
        self.assertEqual(coerce("attendance", "did not attend"), ("No-show", None))
        value, error = coerce("attendance", "probably")
        self.assertIsNone(value)
        self.assertIn("probably", error)

    def test_a_delegate_count_of_zero_stays_zero(self):
        """
        max(1, int(...)) rewrote 4,636 zeros as ones. Whatever a zero means in the
        source it does not mean one, and nothing recorded the change.
        """
        self.assertEqual(coerce("delegate_count", 0), (0, None))
        self.assertEqual(coerce("delegate_count", "0"), (0, None))
        post_import(self.admin, [row(delegate_count=0)])
        self.assertEqual(BookDelegate.objects.get().delegate_count, 0)

    def test_the_old_floor_would_have_rewritten_it(self):
        """Records the defect."""
        self.assertEqual(max(1, int("0")), 1)

    def test_currency_is_not_silently_read_as_usd(self):
        value, error = coerce("currency", "Dollars")
        self.assertIsNone(value)
        self.assertIn("Dollars", error)
        # A file that states nothing still gets the model's own default.
        self.assertIs(coerce("currency", "")[0], UNSET)
        post_import(self.admin, [row()])
        self.assertEqual(BookEvent.objects.get().currency, BookEvent.Currency.USD)

    def test_an_unreadable_date_is_reported_rather_than_stored_as_blank(self):
        """
        The old six-format parser returned None on failure, so a column of
        unreadable dates was indistinguishable from a column of blanks.
        """
        data = post_import(self.admin, [row(payment_date="not a date")])
        self.assertIn("Payment Date", data["errors"][0]["message"])
        self.assertEqual(BookEvent.objects.count(), 0)

    def test_an_excel_serial_in_the_edition_column_is_still_rejected(self):
        data = post_import(self.admin, [row(edition=45678)])
        self.assertIn("Edition", data["errors"][0]["message"])


# ══ 5 PERCENT-AWARE DISCOUNTS ═══════════════════════════════════════════════

class PercentDiscountTests(ImportCase):
    """
    The Discount column mixes "20%" and "0.2" for the same fact. Decimal("20%")
    raises, the bare handler substituted 0.00, and 671 rows imported as a zero
    discount while the 262 decimal-formatted rows imported correctly.
    """

    def test_both_spellings_import_to_the_same_stored_discount(self):
        """The review's acceptance test for fix 5."""
        post_import(self.admin, [
            row(invoice_number="INV-PC", contact_email="pc@example.test", discount="20%"),
            row(invoice_number="INV-FR", contact_email="fr@example.test", discount="0.2"),
        ])
        stored = set(BookDelegate.objects.values_list("discount", flat=True))
        self.assertEqual(stored, {Decimal("0.2000")})

    def test_a_junk_value_is_reported_rather_than_becoming_zero(self):
        data = post_import(self.admin, [row(discount="ask Steve")])
        self.assertIn("Discount", data["errors"][0]["message"])
        self.assertEqual(BookDelegate.objects.count(), 0)

    def test_the_old_handler_would_have_written_zero(self):
        """Records the defect."""
        from decimal import InvalidOperation
        with self.assertRaises(InvalidOperation):
            Decimal("20%")

    def test_it_matches_the_converter_the_browser_already_uses(self):
        """
        percentToFraction in frontend/src/api/bookings.js rounds to four places,
        so the same value typed into the booking form and imported from a file
        must store the same number.
        """
        for raw, expected in (
            ("20%", "0.2000"), ("20", "0.2000"), (20, "0.2000"),
            ("0.2", "0.2000"), ("0.25", "0.2500"), ("25%", "0.2500"),
            ("0", "0.0000"), ("100%", "1.0000"), ("1", "1.0000"),
        ):
            with self.subTest(raw=raw):
                self.assertEqual(percent_to_fraction(raw)[0], Decimal(expected))

    def test_a_discount_outside_0_to_100_percent_is_refused(self):
        self.assertIsNotNone(percent_to_fraction("200%")[1])
        self.assertIsNotNone(percent_to_fraction("-10%")[1])


# ══ 6 BATCH IDENTIFIER ══════════════════════════════════════════════════════

class BatchIdentifierTests(ImportCase):
    """
    The import endpoint stamped no batch identifier and wrote no audit record,
    unlike the Zoho loader which does both. Nothing in the database marked a row
    as belonging to the 26 August import, and the invoice timestamps could not
    stand in because the importer BACKDATES them from an Added Time column.
    """

    def test_an_imports_rows_can_all_be_listed_from_its_batch_id_alone(self):
        """The review's acceptance test for fix 6."""
        data = post_import(self.admin, [
            row(invoice_number="INV-1", contact_email="a@example.test", contact_name="A One"),
            row(invoice_number="INV-2", contact_email="b@example.test", contact_name="B Two"),
        ])
        batch = data["import_batch_id"]
        self.assertTrue(batch)

        self.assertEqual(BookEvent.objects.filter(import_batch_id=batch).count(), 2)
        self.assertEqual(BookDelegate.objects.filter(import_batch_id=batch).count(), 2)
        # And nothing outside the import carries it.
        self.assertEqual(BookEvent.objects.exclude(import_batch_id=batch).count(), 0)

    def test_the_caller_can_pin_one_id_across_every_chunk(self):
        """
        A 20,000-row file is forty sequential calls. Without a client-supplied id
        each chunk would get its own and the import would not be one thing.
        """
        pinned = "11111111-2222-4333-8444-555555555555"
        post_import(self.admin, [row(invoice_number="INV-A", contact_email="a@example.test")],
                    import_batch_id=pinned)
        post_import(self.admin, [row(invoice_number="INV-B", contact_email="b@example.test")],
                    import_batch_id=pinned, batch_number=2)
        self.assertEqual(BookEvent.objects.filter(import_batch_id=pinned).count(), 2)

    def test_a_malformed_batch_id_is_refused_rather_than_ignored(self):
        view = BookEventViewSet.as_view({"post": "bulk_import"})
        request = APIRequestFactory().post(
            "/", {"rows": [row()], "import_batch_id": "not-a-uuid"}, format="json")
        force_authenticate(request, user=self.admin)
        response = view(request)
        self.assertEqual(response.status_code, 400)
        self.assertEqual(BookEvent.objects.count(), 0)

    def test_one_audit_record_is_written_per_call(self):
        from accounts.models import ActionLog

        post_import(self.admin, [row()])
        log = ActionLog.objects.get(action="IMPORTED BOOKINGS")
        self.assertEqual(log.user, self.admin)
        self.assertIn("batch_id=", log.details)
        self.assertIn("inserted=1", log.details)


# ══ 7 DRY RUN ═══════════════════════════════════════════════════════════════

class DryRunTests(ImportCase):
    """
    The review step showed "15,180 rows, 21 columns mapped" and nothing else. A
    dry run would have read "Payable/Free, 11,210 of 15,180 values not
    recognised" while the import could still be cancelled. The tickets importer
    already supported one; bookings did not.
    """

    ROWS = [
        row(invoice_number="INV-1", contact_email="a@example.test", paid_or_free="Payable"),
        row(invoice_number="INV-2", contact_email="b@example.test", paid_or_free="Payable"),
        row(invoice_number="INV-3", contact_email="c@example.test", paid_or_free="Free"),
        row(invoice_number="INV-4", contact_email="d@example.test", paid_or_free="Sponsored"),
        row(invoice_number="INV-5", contact_email="e@example.test", paid_or_free=""),
    ]

    def test_the_review_step_gets_per_column_counts(self):
        """The review's acceptance test for fix 7."""
        data = post_import(self.admin, self.ROWS, dry_run=True)
        self.assertTrue(data["dry_run"])
        self.assertEqual(data["rows"], 5)

        pof = next(c for c in data["columns"] if c["field"] == "paid_or_free")
        self.assertEqual(pof["label"], "Payable / Free")
        self.assertEqual(pof["accepted"], 3)   # two Payable, one Free
        self.assertEqual(pof["blank"], 1)
        self.assertEqual(pof["rejected"], 1)
        self.assertEqual(pof["examples"], [{"value": "Sponsored", "rows": 1}])
        self.assertEqual(sorted(pof["allowed"]), ["Free", "Paid"])

    def test_the_preview_writes_nothing(self):
        post_import(self.admin, self.ROWS, dry_run=True)
        self.assertEqual(BookEvent.objects.count(), 0)
        self.assertEqual(BookDelegate.objects.count(), 0)

    def test_the_rows_that_would_fail_are_named(self):
        data = post_import(self.admin, self.ROWS, dry_run=True)
        self.assertEqual(data["rows_with_errors"], 1)
        self.assertEqual(data["errors"][0]["row_index"], 3)
        self.assertEqual(data["errors"][0]["invoice_number"], "INV-4")

    def test_the_unrecognised_count_is_what_would_have_stopped_the_import(self):
        """
        Shaped like the reviewed file: an overwhelming majority of one column
        unrecognised, which the old review step reported as a clean import.
        """
        rows = [row(invoice_number=f"INV-{i}", contact_email=f"{i}@example.test",
                    paid_or_free="Payable" if i else "Free")
                for i in range(20)]
        # With the alias in place they all resolve, which is the point.
        data = post_import(self.admin, rows, dry_run=True)
        pof = next(c for c in data["columns"] if c["field"] == "paid_or_free")
        self.assertEqual(pof["rejected"], 0)
        self.assertEqual(pof["accepted"], 20)

    def test_column_report_needs_no_database(self):
        """
        The preview is answered from the rows alone, so it cannot write and does
        not need a transaction to prove it did not.
        """
        report = column_report([{"paid_or_free": "Payable"}, {"paid_or_free": "nope"}])
        pof = next(c for c in report if c["field"] == "paid_or_free")
        self.assertEqual((pof["accepted"], pof["rejected"]), (1, 1))


# ══ 8 ONE SHARED COERCION TABLE ═════════════════════════════════════════════

class SharedCoercionTests(TestCase):
    """
    Six paths wrote bookings and each coerced differently. This is the change
    that stops the NEXT variant of the bug rather than this instance of it.
    """

    # Every constrained column on the two models, and the table that must hold
    # an entry for it.
    CONSTRAINED = {
        BookEvent: ("payment_status", "payment_type", "paid_or_free",
                    "ticket_tier", "currency"),
        BookDelegate: ("attendance",),
    }

    def test_every_constrained_column_has_an_entry(self):
        """
        The drift guard. A new choice-validated column cannot be added without
        deciding how it is coerced, which is what "declared per field, with
        allowed values read from the model so it cannot drift" has to mean.
        """
        for model, fields in self.CONSTRAINED.items():
            for name in fields:
                with self.subTest(model=model.__name__, field=name):
                    self.assertIn(name, RULES)

    def test_the_declared_columns_really_are_the_choice_columns(self):
        """
        Guards the guard: if someone adds a sixth choices= column to BookEvent,
        CONSTRAINED above must grow too, or the test above proves nothing.
        """
        for model, expected in self.CONSTRAINED.items():
            actual = {
                f.name for f in model._meta.get_fields()
                if getattr(f, "choices", None) and f.name not in (
                    # Not value-coerced on import: source is set by the write path
                    # itself, and delegate_count is an integer flag with its own
                    # bounded rule rather than a spelling to look up.
                    "source", "delegate_count",
                )
            }
            self.assertEqual(actual, set(expected), f"{model.__name__} choice columns moved")

    def test_allowed_values_are_read_from_the_model(self):
        self.assertEqual(allowed_values("paid_or_free"),
                         sorted(BookEvent.PaidOrFree.values))
        self.assertEqual(allowed_values("ticket_tier"),
                         sorted(BookEvent.TicketTier.values))
        self.assertEqual(allowed_values("attendance"),
                         sorted(BookDelegate.Attendance.values))

    def test_no_alias_can_shadow_a_stored_value(self):
        """
        An alias is applied only where the model declares no such spelling, and
        an alias pointing at a value the model does not declare is a hard error
        at import time rather than a silent no-op.
        """
        for field in ("paid_or_free", "payment_status", "attendance"):
            for stored in allowed_values(field):
                with self.subTest(field=field, stored=stored):
                    self.assertEqual(coerce(field, stored), (stored, None))

    def test_every_write_path_coerces_a_value_identically(self):
        """
        The review's acceptance test for fix 8: the browser import, the two
        commands, the sheet sync and the webhook all coerce a given value
        identically, proven by one shared test.

        Each path is called through its own entry point with the same cell, so
        this fails if any one of them grows a private rule again.
        """
        CASES = (
            ("Payable", "Paid"),
            ("free", "Free"),
            ("FOC", "Free"),
            ("", ""),
            ("Sponsored", ""),   # refused everywhere, never guessed at
        )

        for raw, expected in CASES:
            with self.subTest(raw=raw):
                # 1 the browser import, via the shared table it now calls
                value, error = coerce("paid_or_free", raw)
                browser = "" if (error or value is UNSET) else value

                # 2 import_booking_excel
                from book_event.management.commands.import_booking_excel import (
                    _coerced, _value_warnings,
                )
                _value_warnings.clear()
                excel = _coerced({"Paid/Free": raw}, "Paid/Free", "paid_or_free")

                # 3 sync_bookings_from_sheets
                from book_event.management.commands.sync_bookings_from_sheets import (
                    _build_defaults,
                )
                defaults = _build_defaults(
                    {"paid_or_free": raw}, {"paid_or_free"}, [], [])
                sheets = defaults.get("paid_or_free", "")

                # 4 the website webhook
                from webhooks.services import WebhookProcessor
                processor = WebhookProcessor.__new__(WebhookProcessor)
                processor.notes = []
                webhook = processor._coerce_paid_or_free(raw, "test")

                # 5 the repair command, whose vocabulary this table adopted
                from book_event.management.commands.update_delegate_number_paid_free import (
                    PAID_OR_FREE_LOOKUP,
                )
                repair = PAID_OR_FREE_LOOKUP.get(str(raw).strip().lower(), "")

                self.assertEqual(browser, expected)
                self.assertEqual(excel, expected)
                self.assertEqual(sheets, expected)
                self.assertEqual(webhook, expected)
                self.assertEqual(repair, expected)

    def test_the_webhook_no_longer_defaults_payable_free_to_paid(self):
        """
        F8. A blank field, or any spelling the two-value map did not hold, was
        stored as CHARGED — and written as a per-delegate override, which takes
        precedence over the invoice at read time, so it also overruled anything
        an import wrote. It is why free bookings read as Payable even where the
        source file was correct.
        """
        from webhooks.services import WebhookProcessor

        processor = WebhookProcessor.__new__(WebhookProcessor)
        processor.notes = []
        self.assertEqual(processor._coerce_paid_or_free("", "invoice"), "")
        self.assertEqual(processor._coerce_paid_or_free("Payable", "invoice"), "Paid")

        self.assertEqual(processor._coerce_paid_or_free("Sponsored", "invoice"), "")
        self.assertTrue(
            any("not recognised" in n for n in processor.notes),
            "an unrecognised value must be reported, not silently blanked",
        )

    def test_the_old_webhook_default_is_recorded(self):
        """Records the defect: the fallback was the literal string "Paid"."""
        pof_map = {v.lower(): v for v in BookEvent.PaidOrFree.values}
        self.assertEqual(pof_map.get("".strip().lower(), "Paid"), "Paid")
        self.assertEqual(pof_map.get("payable", "Paid"), "Paid")

    def test_coerce_row_omits_what_the_row_did_not_state(self):
        """
        UNSET is what stops a blank cell overwriting a stored value with a
        default on the upsert path — the behaviour that made the outcome depend
        on whether the invoice already existed.
        """
        values, errors = coerce_row({"paid_or_free": "Payable", "currency": ""})
        self.assertEqual(errors, [])
        self.assertEqual(values, {"paid_or_free": "Paid"})
        self.assertNotIn("currency", values)


# == END TO END, THE QA FIXTURE =============================================

class QaFixtureEndToEndTests(ImportCase):
    """
    Imports `data_imports/qa_import_fixture.xlsx` through the real endpoint and
    asserts every outcome the fixture's own EXPECTED block promises.

    WHY THIS EXISTS ALONGSIDE THE CLASSES ABOVE
    Those build their rows as dicts, which skips the two steps a person actually
    performs: reading a spreadsheet, and letting autoMap resolve its headers.
    This one starts from the real file with the real header row, so it covers the
    whole path -- parse, map, coerce, write, reconcile -- in one assertion set.

    It also keeps the QA document honest. The figures printed by
    scripts/make_qa_import_fixture.py are asserted here, so a change that makes
    the wizard behave differently from the QA plan fails a test rather than
    wasting a tester's afternoon.
    """

    @classmethod
    def fixture_rows(cls):
        """The fixture as bulk_import receives it, headers resolved by autoMap."""
        from pathlib import Path

        import openpyxl
        from django.conf import settings

        path = Path(settings.BASE_DIR) / "data_imports" / "qa_import_fixture.xlsx"
        if not path.exists():
            return None
        ws = openpyxl.load_workbook(path, read_only=True, data_only=True).worksheets[0]
        it = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(it)]
        mapping = {h: HeaderAliasTests.auto_map(h) for h in headers}
        rows = []
        for raw in it:
            if all(c is None or str(c).strip() == "" for c in raw):
                continue
            cells = dict(zip(headers, raw))
            rows.append({t: cells.get(h) for h, t in mapping.items() if t})
        return headers, mapping, rows

    def setUp(self):
        loaded = self.fixture_rows()
        if loaded is None:
            self.skipTest(
                "data_imports/qa_import_fixture.xlsx not present -- "
                "run scripts/make_qa_import_fixture.py"
            )
        self.headers, self.mapping, self.rows = loaded

    # -- the mapping step ---------------------------------------------------
    def test_only_the_deliberate_decoy_column_fails_to_map(self):
        unmapped = [h for h, t in self.mapping.items() if not t]
        self.assertEqual(unmapped, ["Sponsor Lanyard Colour"])
        self.assertEqual(len(self.headers), 22)

    # -- the review step ----------------------------------------------------
    def test_the_review_step_counts_match_the_qa_document(self):
        data = post_import(self.admin, self.rows, dry_run=True)
        got = {c["label"]: (c["accepted"], c["blank"], c["rejected"])
               for c in data["columns"]}
        self.assertEqual(got["Payable / Free"], (9, 0, 1))
        self.assertEqual(got["Ticket Tier"],    (2, 7, 1))
        self.assertEqual(got["Discount"],       (2, 7, 1))
        self.assertEqual(got["Delegate Count"], (10, 0, 0))
        self.assertEqual(got["Attendance"],     (2, 8, 0))
        self.assertEqual(got["Payment Status"], (10, 0, 0))
        self.assertEqual(got["Booking Code"],   (2, 8, 0))
        self.assertEqual(data["rows"], 10)
        self.assertEqual(data["rows_with_errors"], 3)

    def test_the_fixture_preview_writes_nothing(self):
        post_import(self.admin, self.rows, dry_run=True)
        self.assertEqual(BookEvent.objects.count(), 0)
        self.assertEqual(BookDelegate.objects.count(), 0)

    # -- the write ----------------------------------------------------------
    def test_seven_import_and_three_are_reported(self):
        data = post_import(self.admin, self.rows)
        self.assertEqual(data["inserted"], 7)
        self.assertEqual(len(data["errors"]), 3)
        reported = {e["invoice_number"] for e in data["errors"]}
        self.assertEqual(reported, {"QA-BAD-1", "QA-BAD-2", "QA-BAD-3"})
        # Not even partially written.
        self.assertFalse(
            BookEvent.objects.filter(invoice_number__startswith="QA-BAD").exists())
        self.assertFalse(BookDelegate.objects.filter(email__startswith="qa.bad").exists())

    def test_the_mixed_invoice_keeps_both_delegates_values(self):
        post_import(self.admin, self.rows)
        from book_delegate.serializers import BookDelegateInlineSerializer

        rows = {
            d["email"]: d for d in BookDelegateInlineSerializer(
                BookDelegate.objects.filter(invoice_id="QA-MIX-1")
                .select_related("invoice"), many=True).data
        }
        ada, alan = rows["qa.ada@example.test"], rows["qa.alan@example.test"]
        self.assertEqual(
            (ada["effective_paid_or_free"], alan["effective_paid_or_free"]),
            ("Paid", "Free"))
        self.assertEqual((ada["booking_code"], alan["booking_code"]),
                         ("Speaker", "Delegate"))
        self.assertEqual(
            (ada["effective_payment_type"], alan["effective_payment_type"]),
            ("Bank", "Stripe"))
        self.assertEqual(
            (ada["effective_ticket_tier"], alan["effective_ticket_tier"]),
            ("SEB", "EB"))
        self.assertEqual(
            (ada["effective_payment_date"], alan["effective_payment_date"]),
            ("2026-08-03", "2026-08-04"))

    def test_both_discount_spellings_land_on_the_same_number(self):
        post_import(self.admin, self.rows)
        grace = BookDelegate.objects.get(email="qa.grace@example.test")
        katherine = BookDelegate.objects.get(email="qa.katherine@example.test")
        self.assertEqual(grace.discount, katherine.discount)
        self.assertEqual(grace.discount, Decimal("0.20"))

    def test_the_stated_zero_survives_and_attendance_is_translated(self):
        post_import(self.admin, self.rows)
        self.assertEqual(
            BookDelegate.objects.get(email="qa.edsger@example.test").delegate_count, 0)
        self.assertEqual(
            BookDelegate.objects.get(email="qa.barbara@example.test").attendance,
            "Pending")
        self.assertEqual(
            BookDelegate.objects.get(email="qa.donald@example.test").attendance,
            "Confirmed")

    def test_every_written_row_is_listable_from_the_batch_reference(self):
        data = post_import(self.admin, self.rows)
        batch = data["import_batch_id"]
        self.assertEqual(BookEvent.objects.filter(import_batch_id=batch).count(), 6)
        self.assertEqual(BookDelegate.objects.filter(import_batch_id=batch).count(), 7)
        self.assertEqual(BookEvent.objects.exclude(import_batch_id=batch).count(), 0)

    def test_re_importing_the_same_file_changes_nothing(self):
        """The review's verification step, on the real file through the real path."""
        DELEGATE_COLS = (
            "email", "delegate_paid_or_free", "booking_code",
            "delegate_payment_type", "delegate_ticket_tier",
            "delegate_payment_date", "discount", "delegate_count", "attendance",
        )
        INVOICE_COLS = (
            "invoice_number", "paid_or_free", "booking_code", "payment_type",
            "ticket_tier", "payment_date", "delegate_count", "contact_name",
        )
        post_import(self.admin, self.rows)
        before = sorted(BookDelegate.objects.values_list(*DELEGATE_COLS))
        invoices_before = sorted(BookEvent.objects.values_list(*INVOICE_COLS))

        post_import(self.admin, self.rows, duplicate_strategy="upsert")

        self.assertEqual(BookDelegate.objects.count(), 7)
        self.assertEqual(BookEvent.objects.count(), 6)
        self.assertEqual(sorted(BookDelegate.objects.values_list(*DELEGATE_COLS)),
                         before)
        self.assertEqual(sorted(BookEvent.objects.values_list(*INVOICE_COLS)),
                         invoices_before)
