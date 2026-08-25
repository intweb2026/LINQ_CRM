"""
book_event/date_backfill.py
───────────────────────────
Fill the booking DATE columns that are empty in production from a Zoho-style
export, and touch nothing else.

WHY THIS IS SEPARATE FROM THE IMPORTERS
`import_remaining_bookings` writes 28 columns and routes every row through
BookEvent.save()/BookDelegate.save(), which re-derives event_name, re-parses
edition out of event_code, canonicalises booking_code and fills a blank accounts
contact. That is right for an import. It is the wrong instrument for production,
where the ONLY thing missing is dates: any of those derivations firing on 6,000
live invoices is a change nobody asked for, and the value it would write is not
in the workbook to be checked against.

So this module names its columns and writes exactly those:

    Request Date  -> BookEvent.request_date
    Invoice Date  -> BookEvent.invoice_date
    Payment Due   -> BookEvent.payment_due_date
    Date Paid     -> BookEvent.payment_date

plus the one column that is DERIVED from two of them, and which would otherwise
be left stale:

    BookDelegate.booked_on = COALESCE(request_date, invoice_date)

The writes are bulk_update/queryset .update() on those columns alone. Nothing
else on the row can move as a side effect, which is the property that makes this
safe to run against production. It mirrors the sanctioned single-column update
BookEvent.save() performs for booked_on (see book_event/models.py) rather than
calling save() 6,000 times.

BLANK-ONLY BY DEFAULT
A date is written only where the stored column IS NULL. A date already in the
database is never moved unless the caller passes overwrite=True, and a blank cell
in the workbook never clears a stored date. Re-running is therefore a no-op.

ONE VALUE PER INVOICE
The workbook is delegate-grained, and these four columns are invoice-level: 159
invoices list more than one Request Date across their delegates, 162 more than
one Payment Due, 305 more than one Date Paid. An invoice can hold one, so the
FIRST row's value wins — which is what every already-populated row in the
database holds — and every such invoice is listed in the result as `varied` so
the disagreement is visible rather than silently resolved.
"""
from __future__ import annotations

from collections import defaultdict
from datetime import date as Date
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from django.db import transaction

# field on BookEvent -> the workbook header it comes from
DATE_COLUMNS: dict[str, str] = {
    "request_date":     "Request Date",
    "invoice_date":     "Invoice Date",
    "payment_due_date": "Payment Due",
    "payment_date":     "Date Paid",
}

# Filling either of these makes BookDelegate.booked_on stale.
BOOKED_ON_SOURCES = ("request_date", "invoice_date")

INVOICE_HEADER = "Invoice Number"


# ── cell readers ──────────────────────────────────────────────────────────────
def _text(value) -> str:
    """Any cell → stripped string with whitespace runs collapsed; '' for blank."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if text.lower() in ("nan", "none", "nat"):
        return ""
    return " ".join(text.split())


def parse_date(value) -> Optional[Date]:
    """
    A workbook cell → date, or None.

    Slashed dates are read DAY-FIRST, matching accounts.import_common and every
    date already stored. pandas would read "03/04/2026" month-first, so the same
    cell would mean two different days depending on which importer loaded it.
    """
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, Date):
        return value
    text = _text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
                "%d %b %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    # An Excel serial that arrived as a bare number rather than a date cell.
    try:
        serial = float(text)
    except ValueError:
        return None
    from accounts.import_common import excel_serial_to_date
    if 1 < serial < 100_000:
        return excel_serial_to_date(serial)
    return None


def read_workbook_dates(path) -> tuple[dict[str, dict[str, Optional[Date]]],
                                       dict[str, list[str]],
                                       list[str]]:
    """
    Read a workbook into {invoice_number: {field: date}}.

    Returns (dates, varied, unreadable):
      dates      one entry per invoice, from its FIRST row
      varied     {invoice_number: [field, ...]} where the invoice's rows disagree
      unreadable the distinct cell values no format matched, so a mangled date
                 column reports itself instead of importing as a column of blanks
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        stream = ws.iter_rows(values_only=True)
        header = [_text(h) for h in next(stream)]
        missing = [h for h in (INVOICE_HEADER, *DATE_COLUMNS.values())
                   if h not in header]
        if missing:
            raise ValueError(
                f"Workbook is missing required column(s): {', '.join(missing)}"
            )
        index = {name: i for i, name in enumerate(header)}

        dates: dict[str, dict[str, Optional[Date]]] = {}
        seen: dict[str, dict[str, set]] = defaultdict(lambda: defaultdict(set))
        unreadable: set[str] = set()

        for raw in stream:
            if not any(v not in (None, "") for v in raw):
                continue
            invoice = _text(raw[index[INVOICE_HEADER]])
            if not invoice:
                continue
            row_dates = {}
            for field, column in DATE_COLUMNS.items():
                cell = raw[index[column]]
                value = parse_date(cell)
                if value is None and cell not in (None, ""):
                    unreadable.add(f"{column}={cell!r}")
                row_dates[field] = value
                seen[invoice][field].add(value)
            dates.setdefault(invoice, row_dates)

        varied = {
            invoice: sorted(f for f, values in fields.items() if len(values) > 1)
            for invoice, fields in seen.items()
            if any(len(values) > 1 for values in fields.values())
        }
        return dates, varied, sorted(unreadable)
    finally:
        wb.close()


# ── the backfill ──────────────────────────────────────────────────────────────
def backfill_booking_dates(
    path=None,
    *,
    dates: Optional[dict[str, dict[str, Optional[Date]]]] = None,
    fields: Iterable[str] = tuple(DATE_COLUMNS),
    overwrite: bool = False,
    dry_run: bool = False,
    batch_size: int = 1000,
) -> dict:
    """
    Fill the empty booking date columns from a workbook. Returns a result dict.

    path       a workbook to read, or pass `dates` directly for a caller that has
               already parsed one (a test, or a CSV/Sheets source).
    fields     which of DATE_COLUMNS to consider; the rest are left alone.
    overwrite  False (default) writes only where the stored column is NULL.
               True replaces a stored date that differs from the workbook's.
    dry_run    compute and report every write, change nothing.

    Result keys:
      filled          {field: count} rows actually written
      conflicts       {field: count} stored non-NULL and different, left as-is
                      (always empty when overwrite=True, since those are written)
      already_correct {field: count}
      blank_in_file   {field: count} nothing in the workbook to write
      invoices        counted: in_file / matched / missing_from_db
      booked_on       delegates whose derived booked_on was recomputed
      varied          {invoice: [field, ...]} the workbook disagrees with itself
      unreadable      distinct unparseable cell values
      changes         [(invoice, field, stored, written), ...] full change list
      missing         invoice numbers in the workbook with no row in the database
    """
    from book_delegate.models import BookDelegate
    from book_event.models import BookEvent

    fields = tuple(f for f in fields if f in DATE_COLUMNS)
    if not fields:
        raise ValueError(f"No known date fields requested; pick from {list(DATE_COLUMNS)}")

    varied: dict[str, list[str]] = {}
    unreadable: list[str] = []
    if dates is None:
        if path is None:
            raise ValueError("Pass either `path` or `dates`.")
        dates, varied, unreadable = read_workbook_dates(path)

    result = {
        "filled":          {f: 0 for f in fields},
        "conflicts":       {f: 0 for f in fields},
        "already_correct": {f: 0 for f in fields},
        "blank_in_file":   {f: 0 for f in fields},
        "invoices": {"in_file": len(dates), "matched": 0, "missing_from_db": 0},
        "booked_on": 0,
        "varied": {i: [f for f in fs if f in fields] for i, fs in varied.items()},
        "unreadable": unreadable,
        "changes": [],
        "missing": [],
        "dry_run": dry_run,
        "overwrite": overwrite,
        "fields": list(fields),
    }
    result["varied"] = {i: fs for i, fs in result["varied"].items() if fs}

    # Only the columns this touches are loaded; `pk` and invoice_number are
    # needed to write and to key. booked_on's two sources come along regardless
    # of `fields`, because recomputing it needs the FINAL value of both.
    load = set(fields) | set(BOOKED_ON_SOURCES)
    stored = {
        row.invoice_number: row
        for row in BookEvent.objects
        .filter(invoice_number__in=list(dates))
        .only("invoice_number", *load)
    }

    to_update: list[BookEvent] = []
    booked_on_groups: dict[Optional[Date], list[str]] = defaultdict(list)

    for invoice_number, row_dates in dates.items():
        invoice = stored.get(invoice_number)
        if invoice is None:
            result["invoices"]["missing_from_db"] += 1
            result["missing"].append(invoice_number)
            continue
        result["invoices"]["matched"] += 1

        touched = False
        for field in fields:
            new = row_dates.get(field)
            old = getattr(invoice, field)
            if new is None:
                result["blank_in_file"][field] += 1
                continue
            if old == new:
                result["already_correct"][field] += 1
                continue
            if old is not None and not overwrite:
                result["conflicts"][field] += 1
                continue
            setattr(invoice, field, new)
            result["filled"][field] += 1
            result["changes"].append((invoice_number, field, old, new))
            touched = True

        if touched:
            to_update.append(invoice)
            if any(f in fields for f in BOOKED_ON_SOURCES):
                booked_on_groups[
                    invoice.request_date or invoice.invoice_date
                ].append(invoice_number)

    if dry_run:
        result["booked_on"] = sum(
            BookDelegate.objects.filter(invoice_id__in=numbers).count()
            for numbers in booked_on_groups.values()
        )
        return result

    with transaction.atomic():
        if to_update:
            BookEvent.objects.bulk_update(to_update, list(fields), batch_size=batch_size)
        # One UPDATE per distinct resulting date rather than one per invoice.
        # Same single-derived-column write BookEvent.save() performs; filtered on
        # invoice_id, the FK's attname, which holds the invoice-number string
        # (invoice_number is the db_column and is not a query name).
        for value, numbers in booked_on_groups.items():
            for start in range(0, len(numbers), batch_size):
                chunk = numbers[start:start + batch_size]
                result["booked_on"] += (
                    BookDelegate.objects
                    .filter(invoice_id__in=chunk)
                    .exclude(booked_on=value)
                    .update(booked_on=value)
                )

    return result
