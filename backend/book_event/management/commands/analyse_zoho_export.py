"""
book_event/management/commands/analyse_zoho_export.py
──────────────────────────────────────────────────────
Read-only analysis of the Zoho export. Writes NOTHING, ever — no model is
imported for writing and there is no transaction.

WHY THIS EXISTS AND WHY IT IS SEPARATE FROM THE LOADER
Four questions have to be answered from the real file before the booking-code
classification lists in settings can be filled in, and before anyone can say
whether the forthcoming Reports attribution feature will work at all:

  1 what `booking_code` actually contains        → --section booking-code
  2 whether attribution names resolve to Users   → --section attribution
  3 what event-code formats are present          → --section event-codes
  4 how many dates are serials / unparseable     → --section dates

Answering them inside `load_zoho_export --dry-run` was the alternative and is
worse: the loader REJECTS rows it cannot handle, so the rows most worth looking
at are exactly the ones its report summarises away. This command classifies
nothing and rejects nothing; it reports distributions, including of the values
that would fail.

THE ATTRIBUTION QUESTION IS THE POINT
A booking names a person. If those names do not resolve to real accounts, a
report that attributes revenue per user reads zero for everyone and looks like an
empty dataset rather than a defect. So the resolution RATE is reported, and every
unresolved value is listed verbatim with its row count — never truncated silently
(see --limit, which says when it truncated).

NO INVENTED FINDINGS
Every number here is computed from the file passed in. Run against synthetic data
it reports synthetic numbers; that proves the script works, not what the real
export contains.
"""
import json
import re
from collections import Counter

from django.core.management.base import BaseCommand, CommandError

from accounts.import_common import as_text, parse_import_date
from accounts.user_resolution import UserResolver
from webhooks.event_code_normalization import normalise_event_code
from webhooks.event_resolver import Outcome, resolve_event_code

SECTIONS = ("booking-code", "attribution", "event-codes", "dates", "all")

# Column candidates per question. Several spellings because the export's exact
# headers are unknown until it arrives; whichever is present is used, and the
# command says which it picked rather than guessing silently.
BOOKING_CODE_COLUMNS = ("Packages", "Booking_code", "booking_code", "Booking Code")

# The four attribution fields named in the brief, with the model field they would
# land on. NOTE: none of these exist on BookEvent today — the only person FK is
# `sales_executive`. That is itself a finding and is stated in the output.
ATTRIBUTION_COLUMNS = {
    "sales":           ("Sales", "sales", "Sales_Executive", "sales_executive"),
    "speaker_sales":   ("Speaker_Sales", "speaker_sales", "Speaker Sales"),
    "telemarketing":   ("Telemarketing", "telemarketing"),
    "market_research": ("Market_Research", "market_research", "Market Research"),
}

EVENT_CODE_COLUMNS = ("Event_Name.Event_Code_with_Year", "Eventcode",
                      "event_code", "Event Code")

DATE_COLUMNS = ("Invoice_Number.Invoice_Date", "Date", "invoice_date",
                "event_date", "Event_Date")

# Shapes reported for event codes. Purely descriptive — no code is rejected here.
_SHAPES = (
    ("CODE - XX",      re.compile(r"^[A-Za-z0-9/]+\s*-\s*[A-Za-z]+$")),
    ("CODE - XX NN",   re.compile(r"^[A-Za-z0-9/]+\s*-\s*[A-Za-z]+\s+\d{2,4}$")),
    ("CODEnn",         re.compile(r"^[A-Za-z/]+\d{2,4}$")),
    ("CODE only",      re.compile(r"^[A-Za-z0-9/]+$")),
)


def _shape_of(code):
    for name, pattern in _SHAPES:
        if pattern.match(code):
            return name
    return "other"


class Command(BaseCommand):
    help = "Read-only analysis of a Zoho export. Writes nothing."

    def add_arguments(self, parser):
        parser.add_argument("export_file")
        parser.add_argument("--section", choices=SECTIONS, default="all")
        parser.add_argument(
            "--limit", type=int, default=None,
            help=("Cap listed values per section. Omit for the FULL untruncated "
                  "list, which is what the booking-code question needs."),
        )
        parser.add_argument("--json", action="store_true",
                            help="Emit machine-readable JSON instead of a report.")

    def handle(self, *args, **options):
        rows = self._read(options["export_file"])
        if not rows:
            raise CommandError("The export contains no rows.")

        self.limit = options["limit"]
        section = options["section"]
        result = {"rows": len(rows), "columns": sorted(rows[0].keys())}

        if section in ("booking-code", "all"):
            result["booking_code"] = self._booking_code(rows)
        if section in ("attribution", "all"):
            result["attribution"] = self._attribution(rows)
        if section in ("event-codes", "all"):
            result["event_codes"] = self._event_codes(rows)
        if section in ("dates", "all"):
            result["dates"] = self._dates(rows)

        if options["json"]:
            self.stdout.write(json.dumps(result, indent=2, default=str))
        else:
            self.stdout.write(self._render(result))

    # ── input ────────────────────────────────────────────────────────────────
    def _read(self, path):
        lowered = str(path).lower()
        try:
            if lowered.endswith(".json"):
                with open(path, "r", encoding="utf-8") as handle:
                    data = json.load(handle)
                if isinstance(data, dict) and "Event_Bookings_Report" in data:
                    data = data["Event_Bookings_Report"]
                return data if isinstance(data, list) else []
            if lowered.endswith((".xlsx", ".xlsm")):
                from openpyxl import load_workbook
                book = load_workbook(path, data_only=True, read_only=True)
                sheet = book[book.sheetnames[0]]
                stream = sheet.iter_rows(values_only=True)
                header = [as_text(h) for h in next(stream)]
                return [
                    {header[i]: values[i]
                     for i in range(min(len(header), len(values)))}
                    for values in stream
                    if values and not all(v is None for v in values)
                ]
            import csv
            with open(path, "r", encoding="utf-8-sig", newline="") as handle:
                return list(csv.DictReader(handle))
        except FileNotFoundError:
            raise CommandError(f"File not found: {path}")

    @staticmethod
    def _pick(rows, candidates):
        """First candidate column actually present. Reported, never guessed at."""
        present = set(rows[0].keys())
        for name in candidates:
            if name in present:
                return name
        return None

    # ── 1. booking code ──────────────────────────────────────────────────────
    def _booking_code(self, rows):
        column = self._pick(rows, BOOKING_CODE_COLUMNS)
        if column is None:
            return {"column": None,
                    "note": f"none of {list(BOOKING_CODE_COLUMNS)} present"}

        values = [as_text(r.get(column)) for r in rows]
        non_empty = [v for v in values if v]
        counts = Counter(non_empty)
        return {
            "column": column,
            "total_rows": len(values),
            "non_empty": len(non_empty),
            "empty": len(values) - len(non_empty),
            "distinct": len(counts),
            # Full list unless --limit was passed; `truncated` says which.
            "truncated": self.limit is not None and len(counts) > self.limit,
            "values": [{"value": v, "rows": n}
                       for v, n in counts.most_common(self.limit)],
        }

    # ── 2. attribution ───────────────────────────────────────────────────────
    def _attribution(self, rows):
        from book_event.models import BookEvent

        model_fields = {f.name: f for f in BookEvent._meta.get_fields()}
        out = {}

        for logical, candidates in ATTRIBUTION_COLUMNS.items():
            column = self._pick(rows, candidates)
            field = model_fields.get(logical) or model_fields.get("sales_executive"
                                                                  if logical == "sales" else logical)

            entry = {
                "column": column,
                "model_field": getattr(field, "name", None),
                "model_field_kind": (
                    "FK to User" if getattr(field, "many_to_one", False)
                    else type(field).__name__ if field is not None
                    else "ABSENT — no such field on BookEvent"
                ),
            }

            if column is None:
                entry["note"] = f"none of {list(candidates)} present in the file"
                out[logical] = entry
                continue

            resolver = UserResolver()
            values = [as_text(r.get(column)) for r in rows]
            non_empty = [v for v in values if v]
            for value in non_empty:
                resolver.resolve(value)

            entry.update({
                "total_rows": len(values),
                "non_empty": len(non_empty),
                "empty": len(values) - len(non_empty),
                **resolver.report(limit=self.limit),
            })
            out[logical] = entry

        return out

    # ── 3. event codes ───────────────────────────────────────────────────────
    def _event_codes(self, rows):
        from events.models import Event

        column = self._pick(rows, EVENT_CODE_COLUMNS)
        if column is None:
            return {"column": None,
                    "note": f"none of {list(EVENT_CODE_COLUMNS)} present"}

        counts = Counter(v for v in (as_text(r.get(column)) for r in rows) if v)
        catalogue = Event.objects.all()

        shapes = Counter()
        outcomes = Counter()
        unresolved = []

        for code, rows_with_code in counts.items():
            shapes[_shape_of(code)] += rows_with_code
            resolution = resolve_event_code(code, normalise_event_code(code),
                                            queryset=catalogue)
            outcomes[resolution.outcome.value] += rows_with_code
            if resolution.outcome in (Outcome.NO_MATCH, Outcome.AMBIGUOUS):
                unresolved.append({
                    "code": code, "rows": rows_with_code,
                    "outcome": resolution.outcome.value,
                    "candidates": resolution.candidates[:10],
                })

        resolvable = sum(
            n for o, n in outcomes.items()
            if o in (Outcome.EXACT.value, Outcome.BOUNDARY.value,
                     Outcome.BOOKINGS_OFF.value)
        )
        total = sum(counts.values())
        return {
            "column": column,
            "catalogue_size": catalogue.count(),
            "distinct_codes": len(counts),
            "rows_with_a_code": total,
            "shapes": dict(shapes),
            "outcomes": dict(outcomes),
            "resolvable_rows": resolvable,
            "resolvable_pct": (resolvable / total) if total else None,
            "unresolved": sorted(unresolved, key=lambda e: -e["rows"])[:self.limit],
        }

    # ── 4. dates ─────────────────────────────────────────────────────────────
    def _dates(self, rows):
        out = {}
        for column in DATE_COLUMNS:
            if column not in rows[0]:
                continue
            serial = string = blank = failed = 0
            failures = Counter()
            for record in rows:
                raw = record.get(column)
                if raw is None or as_text(raw) == "":
                    blank += 1
                    continue
                if isinstance(raw, (int, float)) and not isinstance(raw, bool):
                    serial += 1
                else:
                    string += 1
                _, error = parse_import_date(raw)
                if error:
                    failed += 1
                    failures[as_text(raw)] += 1
            out[column] = {
                "serial": serial, "string": string, "blank": blank,
                "failed_to_parse": failed,
                "failing_values": [{"value": v, "rows": n}
                                   for v, n in failures.most_common(self.limit)],
            }
        return out or {"note": f"none of {list(DATE_COLUMNS)} present"}

    # ── output ───────────────────────────────────────────────────────────────
    def _render(self, result):
        lines = ["", f"  ZOHO EXPORT ANALYSIS — {result['rows']} rows", ""]

        bc = result.get("booking_code")
        if bc:
            lines.append("  ── booking_code " + "─" * 46)
            if not bc.get("column"):
                lines.append(f"    {bc['note']}")
            else:
                lines += [
                    f"    column      : {bc['column']}",
                    f"    non-empty   : {bc['non_empty']} of {bc['total_rows']}",
                    f"    distinct    : {bc['distinct']}"
                    + ("  (TRUNCATED — rerun without --limit)" if bc["truncated"] else ""),
                    "    values:",
                ]
                lines += [f"      {e['rows']:>7}  {e['value']!r}" for e in bc["values"]]
            lines.append("")

        attribution = result.get("attribution")
        if attribution:
            lines.append("  ── attribution " + "─" * 47)
            for name, entry in attribution.items():
                lines.append(f"    {name}:")
                lines.append(f"      model field : {entry['model_field']} "
                             f"({entry['model_field_kind']})")
                if not entry.get("column"):
                    lines.append(f"      {entry['note']}")
                    continue
                rate = entry.get("resolution_rate")
                lines.append(f"      file column : {entry['column']}")
                lines.append(f"      non-empty   : {entry['non_empty']} of "
                             f"{entry['total_rows']}")
                lines.append(
                    "      resolves    : "
                    + (f"{rate:.1%} ({entry['resolved']}/{entry['attempted']})"
                       if rate is not None else "n/a — nothing non-empty"))
                for item in entry.get("unresolved_values", []):
                    lines.append(f"        UNRESOLVED {item['value']!r} — "
                                 f"{item['rows']} row(s), {item['reason']}")
            lines.append("")

        ec = result.get("event_codes")
        if ec:
            lines.append("  ── event codes " + "─" * 47)
            if not ec.get("column"):
                lines.append(f"    {ec['note']}")
            else:
                pct = ec["resolvable_pct"]
                lines += [
                    f"    column          : {ec['column']}",
                    f"    catalogue size  : {ec['catalogue_size']}",
                    f"    distinct codes  : {ec['distinct_codes']}",
                    f"    shapes          : {ec['shapes']}",
                    f"    outcomes        : {ec['outcomes']}",
                    f"    resolvable rows : {ec['resolvable_rows']} of "
                    f"{ec['rows_with_a_code']}"
                    + (f" ({pct:.1%})" if pct is not None else ""),
                ]
                for item in ec["unresolved"]:
                    lines.append(f"      UNRESOLVED {item['code']!r} — "
                                 f"{item['rows']} row(s), {item['outcome']}, "
                                 f"candidates={item['candidates']}")
            lines.append("")

        dates = result.get("dates")
        if dates:
            lines.append("  ── dates " + "─" * 53)
            for column, entry in dates.items():
                if column == "note":
                    lines.append(f"    {entry}")
                    continue
                lines.append(
                    f"    {column}: serial={entry['serial']} "
                    f"string={entry['string']} blank={entry['blank']} "
                    f"FAILED={entry['failed_to_parse']}")
                for item in entry["failing_values"]:
                    lines.append(f"      unparseable {item['value']!r} — "
                                 f"{item['rows']} row(s)")
            lines.append("")

        return "\n".join(lines)
