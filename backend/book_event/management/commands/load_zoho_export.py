"""
book_event/management/commands/load_zoho_export.py
───────────────────────────────────────────────────
The one-shot load of the Zoho Creator export into `events`, `book_events` and
`book_delegates`.

WHY A COMMAND AND NOT THE BROWSER ENDPOINTS (assumption A1)
`events/bulk_import/` and `invoices/bulk_import/` stay as they are for ad-hoc
user imports. They are the wrong tool for a 35k-row one-shot load, and each
reason is a requirement below: they have no preview, they wrap EACH ROW in its
own savepoint (so a failure at row 20,000 leaves 19,999 rows written), they stamp
no batch id, they write no ActionLog, their date parser silently returns None on
anything it cannot read, and they never consult the event-code resolver.

THE NINE REQUIREMENTS, AND WHERE EACH LIVES
  1 --dry-run writes nothing            → run(), gated on `self.dry_run`
  2 one transaction for the whole load  → handle(), a single transaction.atomic()
  3 one import_batch_id per load        → self.batch_id, stamped in _write_*
  4 one ActionLog per load              → _write_action_log()
  5 dates via parse_import_date         → _row_dates(), rejects quoting the raw
  6 codes via resolve_event_code        → _resolve_event()
  7 dependency order enforced           → run(): events → invoices → delegates
  8 idempotent                          → _plan_*, keyed as documented below
  9 auto-generated invoice numbers      → _invoice_number(), counted

INPUT SHAPE
One FLAT file, one row per delegate, with invoice and event columns repeated —
the shape `import_bookings_json` already documents for the Zoho "Event Bookings
Report". All three levels are derived from it: distinct event codes, distinct
invoice numbers, then one delegate per row.

IDEMPOTENCY KEYS (requirement 8)
  Event        → event_code            (unique=True)
  BookEvent    → invoice_number        (unique=True)
  BookDelegate → (invoice, email)      (unique_together)
An existing key is SKIPPED and counted, never duplicated and never updated — an
update would silently overwrite whatever a user had corrected since the load.

THE AUTO-INVOICE / IDEMPOTENCY TRAP (requirement 9 vs 8)
A2 says a row with no invoice number gets one generated. `invoices/bulk_import/`
does that with `uuid4()`, which is NOT re-runnable: the second run mints
different numbers, every key misses, and the load duplicates exactly the rows it
was supposed to skip. So the number here is DERIVED from the row's own stable
content (event code, email, name, invoice date) — the same row yields the same
number on every run, and requirement 8 holds for generated numbers too. The
count of generated numbers is reported, never silent.

MAIL
This command sends nothing and imports no mail machinery. The alert that fired
once per 500-row chunk from `invoices/bulk_import/` has no equivalent here; see
accounts/tests_import_alert_suppression.py for the endpoint-side guard.
"""
import hashlib
import json
import uuid
from collections import Counter, defaultdict

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounts.import_common import (
    as_text,
    build_header_mapper,
    normalise_row,
    parse_edition,
    parse_import_date,
)
from book_delegate.models import BookDelegate
from book_event.models import BookEvent
from events.models import Event
from webhooks.event_code_normalization import normalise_event_code
from webhooks.event_resolver import Outcome, resolve_event_code

# ── Zoho header map ──────────────────────────────────────────────────────────
# Keys are clean_header()'d (trimmed, whitespace-collapsed, lower-cased). The
# dotted names are the Zoho Report lookup columns that import_bookings_json.py
# already handles, which is the only first-hand evidence of the export's shape
# available while the file itself is absent. Both the dotted and the plain
# spellings are accepted because that command accepts both.
ZOHO_HEADERS = {
    "invoice_number.invoice_number": "invoice_number",
    "invoicenumber":                 "invoice_number",
    "invoice number":                "invoice_number",
    "event_name.event_code_with_year": "event_code",
    "eventcode":                     "event_code",
    "event code":                    "event_code",
    "event_name":                    "event_name",
    "eventname":                     "event_name",
    "invoice_number.invoice_date":   "invoice_date",
    "invoice_date":                  "invoice_date",
    "date":                          "invoice_date",
    "event_date":                    "event_date",
    "sub_company":                   "company_name",
    "delegatecompanyname":           "company_name",
    "company_name":                  "company_name",
    "account_emails":                "accounts_contact_email",
    "accountscontactemail":          "accounts_contact_email",
    "status":                        "payment_status",
    "paymentstatus":                 "payment_status",
    "payment_type":                  "payment_type",
    "paymenttype":                   "payment_type",
    "booking_code_type":             "paid_or_free",
    "paidorfree":                    "paid_or_free",
    "ticket_tier":                   "ticket_tier",
    "tickettier":                    "ticket_tier",
    "packages":                      "booking_code",
    "booking_code":                  "booking_code",
    "discount":                      "discount",
    "edition":                       "edition",
    "name":                          "contact_name",
    "firstname":                     "first_name",
    "lastname":                      "last_name",
    "delegate_email":                "email",
    "email":                         "email",
    "direct_line":                   "phone_number",
    "phonenumber":                   "phone_number",
    "sales_executive":               "sales_executive",
    "sales":                         "sales_executive",
}

MODEL_FIELDS = {
    "invoice_number", "event_code", "event_name", "invoice_date", "event_date",
    "company_name", "accounts_contact_email", "payment_status", "payment_type",
    "paid_or_free", "ticket_tier", "booking_code", "discount", "edition",
    "contact_name", "first_name", "last_name", "email", "phone_number",
    "sales_executive",
}

DATE_FIELDS = ("invoice_date", "event_date")

CREATE, SKIP, REJECT = "create", "skip", "reject"


def _split_name(full):
    parts = as_text(full).split(" ", 1)
    if not parts or not parts[0]:
        return "", ""
    return parts[0], (parts[1] if len(parts) > 1 else "")


class Command(BaseCommand):
    help = "Load the Zoho Creator export into events, book_events and book_delegates."

    def add_arguments(self, parser):
        parser.add_argument("export_file", help="Path to the Zoho export (.xlsx/.csv/.json)")
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Validate and report what WOULD happen. Writes nothing.",
        )
        parser.add_argument(
            "--batch-id", default=None,
            help="Reuse a specific import_batch_id (UUID). Defaults to a new one.",
        )
        parser.add_argument(
            "--operator", default=None,
            help="Username to attribute the ActionLog to. Defaults to the first admin.",
        )
        parser.add_argument(
            "--limit", type=int, default=None,
            help="Process only the first N rows. For rehearsal on a scratch DB.",
        )
        parser.add_argument(
            "--create-missing-events", action="store_true",
            help=("Create catalogue entries for event codes not already in "
                  "`events`. OFF by default — see the note in _stage_events."),
        )
        # Test hook for requirement 2. Named for what it is; there is no way to
        # trip it accidentally from the command line in normal use.
        parser.add_argument(
            "--fail-after", type=int, default=None,
            help="Abort deliberately after N delegate writes, to prove atomicity.",
        )

    # ── entry ────────────────────────────────────────────────────────────────
    def handle(self, *args, **options):
        self.dry_run = options["dry_run"]
        self.fail_after = options["fail_after"]
        self._operator = options["operator"]
        self.create_missing_events = options["create_missing_events"]
        self.batch_id = uuid.UUID(options["batch_id"]) if options["batch_id"] else uuid.uuid4()

        rows = self._read(options["export_file"], options["limit"])
        if not rows:
            raise CommandError("The export contains no rows.")

        # Columns from the UNION of every row, not just the first. A JSON export
        # is a list of objects with no schema, and Zoho omits keys whose value is
        # blank — so a column absent from row 1 but present on row 900 would
        # otherwise never be mapped, and would be silently dropped for the whole
        # file. Found by a synthetic row carrying an `edition` column that row 1
        # did not have: the bad edition sailed through unvalidated.
        # dict.fromkeys preserves first-seen order, so the report lists
        # unrecognised columns in file order rather than an arbitrary set order.
        columns = list(dict.fromkeys(key for record in rows for key in record))
        mapping, unrecognised = build_header_mapper(ZOHO_HEADERS, MODEL_FIELDS)(
            columns)
        if not mapping:
            raise CommandError(
                "No recognisable columns. Got: " + ", ".join(map(str, rows[0].keys())))

        self.stats = Counter()
        self.rejections = []          # [{row, stage, field, problem, value}]
        self._rejected_keys = set()   # (row, field, problem) — see _reject
        # Keyed by row index, not a list: _invoice_number is called once by the
        # invoice stage and again by the delegate stage, and appending on both
        # would report twice as many generated numbers as rows.
        self.generated_invoices = {}  # row index -> invoice_number
        self.unrecognised = unrecognised

        normalised = [normalise_row(r, mapping) for r in rows]

        # Requirement 2: ONE transaction for the entire load. A failure anywhere —
        # including the deliberate --fail-after — rolls back every stage, not just
        # the one that raised. --dry-run additionally rolls back unconditionally,
        # so even a coding error in a _write_ path cannot leave rows behind.
        try:
            with transaction.atomic():
                result = self.run(normalised)
                if self.dry_run:
                    raise _DryRun(result)
        except _DryRun as done:
            result = done.result
        return self._report(result)

    # ── file reading ─────────────────────────────────────────────────────────
    def _read(self, path, limit):
        lowered = str(path).lower()
        try:
            if lowered.endswith(".json"):
                rows = self._read_json(path)
            elif lowered.endswith((".xlsx", ".xlsm")):
                rows = self._read_xlsx(path)
            else:
                rows = self._read_csv(path)
        except FileNotFoundError:
            raise CommandError(f"File not found: {path}")
        return rows[:limit] if limit else rows

    def _read_json(self, path):
        with open(path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
        # Same wrapper import_bookings_json accepts.
        if isinstance(data, dict) and "Event_Bookings_Report" in data:
            data = data["Event_Bookings_Report"]
        if not isinstance(data, list):
            raise CommandError(
                "Expected a list or {'Event_Bookings_Report': [...]}.")
        return data

    def _read_csv(self, path):
        import csv
        with open(path, "r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))

    def _read_xlsx(self, path):
        # data_only=True so formula cells yield their cached VALUE. Dates arrive
        # as real datetimes where Excel typed them and as ints where it did not;
        # parse_import_date handles both, which is the entire point of routing
        # every date through it (requirement 5).
        from openpyxl import load_workbook

        book = load_workbook(path, data_only=True, read_only=True)
        sheet = book[book.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            header = [as_text(h) for h in next(rows)]
        except StopIteration:
            return []
        out = []
        for values in rows:
            if values is None or all(v is None or as_text(v) == "" for v in values):
                continue
            out.append({header[i]: values[i] for i in range(min(len(header), len(values)))})
        return out

    # ── helpers ──────────────────────────────────────────────────────────────
    def _reject(self, index, stage, field, problem, value=None):
        """
        Deduplicated on (row, field, problem), NOT on stage. One bad invoice_date
        is seen by both the event stage and the invoice stage — reporting it twice
        would make "rows rejected" count issues rather than rows, and inflate the
        number an operator uses to decide whether the file is usable.
        """
        key = (index, field, problem)
        if key in self._rejected_keys:
            return
        self._rejected_keys.add(key)
        self.rejections.append({
            "row": index + 1, "stage": stage, "field": field,
            "problem": problem, "value": repr(value) if value is not None else "",
        })

    def _row_dates(self, row, index, stage):
        """
        Requirement 5. Every date column through parse_import_date. An
        unparseable value REJECTS the row quoting the raw value — the legacy
        `_parse_date` returned None here, so a whole column of unreadable dates
        looked exactly like a column of blanks.
        """
        out, ok = {}, True
        for field in DATE_FIELDS:
            if field not in row:
                continue
            value, error = parse_import_date(row[field])
            if error:
                self._reject(index, stage, field, error, row[field])
                ok = False
                continue
            out[field] = value
        return out, ok

    def _invoice_number(self, row, index):
        """
        Requirement 9 + A2, without breaking requirement 8. See the module
        docstring: DERIVED, not random, so a re-run reproduces it exactly.
        """
        given = as_text(row.get("invoice_number"))
        if given:
            return given, False

        seed = "|".join([
            as_text(row.get("event_code")).upper(),
            as_text(row.get("email")).lower(),
            as_text(row.get("contact_name")).lower(),
            as_text(row.get("invoice_date")),
        ])
        digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:10].upper()
        generated = f"IMP-{digest}"
        self.generated_invoices[index] = generated
        return generated, True

    def _resolve_event(self, raw_code, index, stage, catalogue):
        """
        Requirement 6. Boundary-anchored resolution through the shared resolver —
        NOT the exact-match-after-stripping that events/views.py and
        book_event/views.py use today, whose failure mode is no match at all.
        """
        raw = as_text(raw_code)
        if not raw:
            self._reject(index, stage, "event_code", "no event code on the row")
            return None
        resolution = resolve_event_code(raw, normalise_event_code(raw),
                                        queryset=catalogue)
        if resolution.ok:
            return resolution.event

        # BOOKINGS_OFF means the anchored rule DID match — every match simply has
        # web bookings switched off. That outcome exists for live webhook
        # ingestion, where "that edition is closed" is the correct answer to a new
        # booking. This is a historical load: the bookings already happened, and
        # most editions being loaded are finished, so refusing them would reject
        # nearly the whole file for a reason that does not apply.
        #
        # Accepted ONLY when the match is unambiguous. AMBIGUOUS and NO_MATCH
        # still reject — those are the failures anchoring exists to catch, and
        # guessing between two editions is the original bug.
        if resolution.outcome is Outcome.BOOKINGS_OFF and len(resolution.matches) == 1:
            return resolution.matches[0]

        self._reject(index, stage, "event_code",
                     resolution.error_message or resolution.outcome.value, raw)
        return None

    # ── the load, in dependency order (requirement 7) ────────────────────────
    def run(self, rows):
        """
        Events, then invoices, then delegates. Enforced here rather than left to
        the operator: an invoice cannot resolve against a catalogue that has not
        been written, and a delegate's FK points at an invoice_number.
        """
        events = self._stage_events(rows)
        invoices = self._stage_invoices(rows)
        delegates = self._stage_delegates(rows, invoices)

        if not self.dry_run:
            self._write_action_log(events, invoices, delegates)

        return {"events": events, "invoices": invoices, "delegates": delegates}

    # ── stage 1: events ──────────────────────────────────────────────────────
    def _stage_events(self, rows):
        """
        Distinct event codes. Keyed on Event.event_code (unique).

        CREATION IS OPT-IN, AND THAT IS THE WHOLE POINT
        `events` is a curated 142-row master catalogue. If this stage created an
        Event for every code it saw, requirement 6 would be dead on arrival: the
        anchored resolver could never reject anything, because the invoice stage
        would always find a row this stage had just invented from the same string.
        A typo would silently become a new event, and the booking would attach to
        it — which is a worse version of the bug event_resolver.py exists to stop,
        since at least a mis-resolution attaches to a REAL event.

        So by default nothing is created: codes are resolved against the existing
        catalogue and an unresolvable code rejects its rows, loudly, with the
        candidates listed. `--create-missing-events` turns creation on for the
        case where the export genuinely introduces new editions — an explicit
        operator decision, made after reading the dry-run's rejection list.
        """
        wanted = {}
        for index, row in enumerate(rows):
            code = as_text(row.get("event_code"))
            if not code:
                continue
            dates, ok = self._row_dates(row, index, "event")
            if not ok:
                continue
            wanted.setdefault(normalise_event_code(code), {
                "event_code": code,
                "event_date": dates.get("event_date") or dates.get("invoice_date"),
                "name": as_text(row.get("event_name")),
            })

        existing = set(
            Event.objects.values_list("event_code", flat=True))
        existing_norm = {normalise_event_code(c) for c in existing}

        created, skipped = [], []
        for key, spec in wanted.items():
            if key in existing_norm:
                skipped.append(spec["event_code"])
                continue
            if not self.create_missing_events:
                # Not an error here — the invoice stage will reject the row and
                # report the resolver's candidate list, which is the more useful
                # message. Counting it twice would double the rejection total.
                continue
            if not self.dry_run:
                Event.objects.create(
                    event_code=spec["event_code"],
                    official_event_name=spec["name"],
                    # Event.event_date is NOT NULL. A code seen only on undated
                    # rows cannot be invented a date, so it is rejected rather
                    # than defaulted to today — which is what events/bulk_import/
                    # does, and is how "Untitled Event" rows appear.
                    event_date=spec["event_date"],
                    import_batch_id=self.batch_id,
                )
            created.append(spec["event_code"])

        # Codes with no usable date cannot be created; surface them as rejects.
        # Only meaningful when we are creating — otherwise the catalogue already
        # holds the event and its date, and the file's dates are irrelevant here.
        undated = ([s["event_code"] for s in wanted.values() if not s["event_date"]]
                   if self.create_missing_events else [])
        for code in undated:
            if code in created:
                created.remove(code)
            self.rejections.append({
                "row": "-", "stage": "event", "field": "event_date",
                "problem": "no parseable date on any row carrying this code",
                "value": repr(code),
            })

        return {"created": created, "skipped": skipped}

    # ── stage 2: invoices ────────────────────────────────────────────────────
    def _stage_invoices(self, rows):
        """Distinct invoice numbers. Keyed on BookEvent.invoice_number (unique)."""
        from accounts.user_resolution import UserResolver

        self.resolver = UserResolver()
        catalogue = Event.objects.all()

        wanted, order = {}, []
        for index, row in enumerate(rows):
            inv_no, generated = self._invoice_number(row, index)
            if inv_no in wanted:
                continue

            event = self._resolve_event(row.get("event_code"), index, "invoice",
                                        catalogue)
            if event is None:
                continue

            dates, ok = self._row_dates(row, index, "invoice")
            if not ok:
                continue

            edition, edition_error = parse_edition(row.get("edition"))
            if edition_error:
                self._reject(index, "invoice", "edition", edition_error,
                             row.get("edition"))
                continue

            sales_exec, _ = self.resolver.resolve(row.get("sales_executive"))

            wanted[inv_no] = {
                "invoice_number": inv_no,
                "event": event,
                "generated": generated,
                "invoice_date": dates.get("invoice_date"),
                "edition": edition,
                "booking_code": as_text(row.get("booking_code")),
                "company_name": as_text(row.get("company_name")),
                "contact_name": as_text(row.get("contact_name")),
                "contact_email": as_text(row.get("email")).lower(),
                "accounts_contact_email": as_text(row.get("accounts_contact_email")),
                "sales_executive": sales_exec,
            }
            order.append(inv_no)

        existing = set(BookEvent.objects.filter(
            invoice_number__in=list(wanted)).values_list("invoice_number", flat=True))

        created, skipped = [], []
        for inv_no in order:
            if inv_no in existing:
                skipped.append(inv_no)
                continue
            spec = wanted[inv_no]
            if not self.dry_run:
                BookEvent.objects.create(
                    invoice_number=spec["invoice_number"],
                    event_code=spec["event"].event_code,
                    event_name=spec["event"].name,
                    invoice_date=spec["invoice_date"],
                    edition=spec["edition"],
                    booking_code=spec["booking_code"],
                    company_name=spec["company_name"],
                    contact_name=spec["contact_name"],
                    contact_email=spec["contact_email"],
                    accounts_contact_email=spec["accounts_contact_email"],
                    sales_executive=spec["sales_executive"],
                    source=BookEvent.Source.MANUAL,
                    import_batch_id=self.batch_id,
                )
            created.append(inv_no)

        self._invoice_specs = wanted
        return {"created": created, "skipped": skipped}

    # ── stage 3: delegates ───────────────────────────────────────────────────
    def _stage_delegates(self, rows, invoices):
        """One per row. Keyed on (invoice, email) — the model's unique_together."""
        writable = set(invoices["created"]) | set(invoices["skipped"])

        existing = set(
            BookDelegate.objects.filter(invoice_id__in=list(writable))
            .values_list("invoice_id", "email")
        )

        created, skipped = 0, 0
        seen_in_file = set()

        for index, row in enumerate(rows):
            inv_no, _ = self._invoice_number(row, index)
            if inv_no not in writable:
                continue   # its invoice was rejected; already counted there

            email = as_text(row.get("email")).lower()
            if not email:
                self._reject(index, "delegate", "email", "no email on the row")
                continue

            key = (inv_no, email)
            # Two identical (invoice, email) rows inside ONE file would violate
            # unique_together at write time; the second is a duplicate of the
            # first, not a second person, so it is skipped like any re-run row.
            if key in existing or key in seen_in_file:
                skipped += 1
                continue
            seen_in_file.add(key)

            first, last = _split_name(row.get("contact_name"))
            if row.get("first_name"):
                first, last = as_text(row.get("first_name")), as_text(row.get("last_name"))
            if not first:
                self._reject(index, "delegate", "first_name", "no name on the row")
                continue

            if not self.dry_run:
                spec = self._invoice_specs.get(inv_no, {})
                BookDelegate.objects.create(
                    invoice_id=inv_no,
                    event_code=spec.get("event").event_code if spec.get("event") else "",
                    edition=spec.get("edition"),
                    first_name=first,
                    last_name=last,
                    email=email,
                    phone_number=as_text(row.get("phone_number")),
                    company_name_raw=as_text(row.get("company_name")),
                    import_batch_id=self.batch_id,
                )
                # Requirement 2's proof hook: blow up mid-stage, having already
                # written events, invoices and some delegates.
                if self.fail_after is not None and created + 1 >= self.fail_after:
                    raise RuntimeError(
                        f"--fail-after {self.fail_after}: deliberate abort to "
                        f"prove the whole load rolls back")
            created += 1

        return {"created": created, "skipped": skipped}

    # ── audit (requirement 4) ────────────────────────────────────────────────
    def _write_action_log(self, events, invoices, delegates):
        from accounts.models import ActionLog, User

        operator = (
            User.objects.filter(username=self._operator).first()
            if getattr(self, "_operator", None) else None
        ) or User.objects.filter(role=User.Role.ADMIN).order_by("id").first()
        if operator is None:
            operator = User.objects.order_by("id").first()
        if operator is None:
            # No users at all: an ActionLog needs a NOT NULL user_id, and a load
            # into an empty database has no operator to name. Say so rather than
            # crash the load over its own audit row.
            self.stderr.write("No user exists to attribute the ActionLog to — skipped.")
            return

        ActionLog.objects.create(
            user=operator,
            action=(f"Zoho load: {len(events['created'])} events, "
                    f"{len(invoices['created'])} invoices, "
                    f"{delegates['created']} delegates"),
            details=(
                f"import_batch_id={self.batch_id}\n"
                f"events created={len(events['created'])} skipped={len(events['skipped'])}\n"
                f"invoices created={len(invoices['created'])} skipped={len(invoices['skipped'])}\n"
                f"delegates created={delegates['created']} skipped={delegates['skipped']}\n"
                f"invoice_numbers_generated={len(self.generated_invoices)}\n"
                f"rejected_rows={len(self.rejections)}\n"
                f"unrecognised_columns={self.unrecognised}\n"
                f"sales_executive_resolution={self.resolver.report(limit=50)}"
            ),
        )

    # ── output ───────────────────────────────────────────────────────────────
    def _report(self, result):
        events, invoices, delegates = (
            result["events"], result["invoices"], result["delegates"])
        mode = "DRY RUN — nothing written" if self.dry_run else "LOAD COMMITTED"

        lines = [
            "",
            f"  {mode}",
            f"  import_batch_id : {self.batch_id}",
            "",
            f"  events     created {len(events['created']):>6}   skipped {len(events['skipped']):>6}",
            f"  invoices   created {len(invoices['created']):>6}   skipped {len(invoices['skipped']):>6}",
            f"  delegates  created {delegates['created']:>6}   skipped {delegates['skipped']:>6}",
            "",
            f"  invoice numbers generated : {len(self.generated_invoices)}",
            f"  rows rejected             : {len(self.rejections)}",
        ]

        if self.unrecognised:
            lines.append(f"  unrecognised columns      : {', '.join(map(str, self.unrecognised))}")

        rate = self.resolver.resolution_rate if hasattr(self, "resolver") else None
        if rate is not None:
            lines.append(
                f"  sales exec resolution     : {rate:.1%} "
                f"({self.resolver.resolved_count}/{self.resolver.attempted_count})")
            if self.resolver.unresolved:
                lines.append("    unresolved values:")
                for entry in self.resolver.report(limit=20)["unresolved_values"]:
                    lines.append(
                        f"      {entry['value']!r} — {entry['rows']} row(s), {entry['reason']}")

        if self.rejections:
            lines.append("")
            lines.append("  REJECTED ROWS (per-row reasons):")
            by_stage = defaultdict(list)
            for entry in self.rejections:
                by_stage[entry["stage"]].append(entry)
            for stage, entries in by_stage.items():
                lines.append(f"    {stage}: {len(entries)}")
                for entry in entries[:25]:
                    value = f" — got {entry['value']}" if entry["value"] else ""
                    lines.append(
                        f"      row {entry['row']}: {entry['field']}: "
                        f"{entry['problem']}{value}")
                if len(entries) > 25:
                    lines.append(f"      … and {len(entries) - 25} more")

        lines.append("")
        self.stdout.write("\n".join(lines))
        return None


class _DryRun(Exception):
    """
    Carries the result out through the rollback. --dry-run must not merely
    "choose not to write": raising inside the atomic block means the database
    rolls the transaction back regardless of what any stage did, so a preview
    cannot leave rows behind even if a _write_ path is later changed carelessly.
    """

    def __init__(self, result):
        super().__init__("dry run")
        self.result = result
