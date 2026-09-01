"""
management command: import_remaining_bookings

Pushes a Zoho-style booking export into the CRM **without wiping anything**.
Unlike `import_booking_excel`, which deletes every BookEvent/BookDelegate and
rebuilds them, this command matches each workbook row onto the row that is
already stored and writes only the columns that actually differ.

WHY A SECOND IMPORTER EXISTS
`remaining data.xlsx` is the same dataset that is already in the database — 6112
of its 6113 invoices and 8110 of its 8111 delegates match a stored row — but it
carries four columns the earlier import never mapped, so those columns are empty
for every historical row:

    Payment Due  -> BookEvent.payment_due_date   (8004 rows blank in the DB)
    Parent Code  -> BookEvent.parent_code        (6446 rows blank)
    Job Title    -> BookDelegate.position        (6033 rows blank)
    Discount     -> BookEvent/BookDelegate.discount  (629 rows still 0.00)

Re-running the wiping importer would fill those, but it would also drop every
edit made in the CRM since, so the work is an UPDATE, not a re-import.

ROW MATCHING
    invoice  : Invoice Number == BookEvent.invoice_number  (exact)
    delegate : (Invoice Number, Delegate Email) == (invoice, email)
               falling back to (Invoice Number, full name), because two rows in
               this workbook share an email with a different person on the same
               invoice and the stored copies of those two carry the synthetic
               `dup-<hash>@import.local` address the first import gave them.
    Anything still unmatched is CREATED.

COLUMN -> FIELD MAP
  Invoice-level (BookEvent)
    Payment Status    -> payment_status
    Event Code        -> event_code          (new invoices only; see below)
    Event Name        -> event_name          (re-derived by BookEvent.save())
    Booking Code      -> booking_code        (canonicalised by save())
    Request Date      -> request_date
    Invoice Date      -> invoice_date
    Payment Due       -> payment_due_date
    Date Paid         -> payment_date
    Payment Type      -> payment_type
    Paid/Free         -> paid_or_free
    Ticket Tier       -> ticket_tier
    Parent Code       -> parent_code
    Discount          -> discount            (0.2 == a 20% discount, as stored)
    Add-Ons           -> add_ons
    Ref               -> reference
    Account Company   -> company_name        (--company-source, default account)
    Accounts Contact  -> accounts_contact_email
    Sales Executive   -> sales_executive     (FK, matched on full name)
    first row's Name / Delegate Email / Direct Line
                      -> contact_name / contact_email / contact_phone
    row count         -> delegate_count

  Delegate-level (BookDelegate)
    Name              -> first_name + last_name
    Job Title         -> position
    Delegate Company  -> company_name_raw
    Delegate Email    -> email
    Direct Line       -> phone_number
    Delegate Number   -> delegate_count      (0/1 flag, --delegate-number-as)
    Attendance - IN?  -> attendance          (TRUE -> Confirmed, else Pending)
    Booking Code      -> booking_code
    Discount          -> discount
    Add-Ons           -> add_ons
    Ref               -> reference

  Event Code is NOT rewritten on an invoice that already exists: the stored code
  is already canonical ("Feb2027_SFIL-AD" + edition 2027) while the workbook
  spells it with the year glued on ("Feb2027_SFIL-AD27"). For a NEW invoice the
  code is resolved by exact match, then by dropping the trailing year, then by
  the master-code + year lookup the other importers use.

CONFLICT POLICY (--conflicts)
    fill-blanks  (default) write only where the stored value is empty. Cannot
                 change a value a human has typed into the CRM.
    file-wins    write wherever the workbook differs from the database.
    Neither mode ever clears a stored value with a blank cell unless
    --allow-clear is passed. Comparisons ignore differences that are pure
    formatting: whitespace runs, "545" vs "545.00", and a phone number's leading
    "+" (the workbook exports it as a bare number; the database stores "+…").

Usage:
    python manage.py import_remaining_bookings "path/to/remaining data.xlsx" --dry-run
    python manage.py import_remaining_bookings "path/to/remaining data.xlsx" \
        --conflicts file-wins --report import_remaining_report.md
"""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import date as Date
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from book_delegate.models import BookDelegate
from book_event.models import BookEvent
from events.models import Event
from event_performance.active_edition_service import (
    extract_year_from_code,
    normalize_master_code,
)

User = get_user_model()

# The 26 headers this workbook carries. A missing one is an error rather than a
# silently skipped column, which is how the four unmapped columns went unnoticed
# through the first import.
REQUIRED_HEADERS = (
    "Payment Status", "Event Code", "Booking Code", "Request Date",
    "Invoice Date", "Payment Due", "Invoice Number", "Name", "Job Title",
    "Delegate Company", "Delegate Email", "Direct Line", "Account Company",
    "Accounts Contact", "Delegate Number", "Paid/Free", "Parent Code",
    "Date Paid", "Payment Type", "Ticket Tier", "Discount", "Add-Ons", "Ref",
    "Event Name", "Sales Executive", "Attendance - IN?",
)


# ── cell readers ──────────────────────────────────────────────────────────────
def _s(v) -> str:
    """Any cell → stripped string; '' for blank. Whitespace runs collapsed."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    return " ".join(s.split())


def _date(v) -> Optional[Date]:
    """A date cell → date. openpyxl already hands back datetime for date cells."""
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, Date):
        return v
    raw = _s(v)
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _dec(v) -> Optional[Decimal]:
    """A numeric cell → Decimal, or None when blank."""
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return None


def _bool(v) -> bool:
    if isinstance(v, bool):
        return v
    return _s(v).lower() in ("true", "1", "yes", "y")


def _phone(v) -> str:
    """
    A Direct Line cell in the spelling the database already uses.

    The workbook exports the number as a bare integer (11831234984) while every
    stored row has the leading "+" (+11831234984). Without this, all 3,850
    populated phone numbers would read as changed and be rewritten to a
    different spelling of the same number.
    """
    raw = _s(v)
    if not raw:
        return ""
    if raw.startswith("+"):
        return raw
    digits = re.sub(r"\D", "", raw)
    return f"+{digits}" if digits and digits == raw else raw


def _split_name(full: str) -> tuple[str, str]:
    parts = full.split(None, 1)
    if not parts:
        return "Unknown", ""
    return parts[0], (parts[1] if len(parts) > 1 else "")


def _name_key(first: str, last: str) -> str:
    return " ".join(f"{first} {last}".split()).lower()


# ── comparison ────────────────────────────────────────────────────────────────
def _same(a, b) -> bool:
    """
    True when two values mean the same thing.

    Formatting-only differences must not count as changes, or a re-run would
    rewrite thousands of rows into a different spelling of what they hold:
      - "Variable Speed Hydraulics  (VSH)" vs "… (VSH)"  (collapsed whitespace)
      - "545" vs "545.00"                                (Add-Ons as a number)
      - Decimal("0.2") vs Decimal("0.20")                (scale)
    """
    if a is None:
        a = ""
    if b is None:
        b = ""
    if isinstance(a, Decimal) or isinstance(b, Decimal):
        try:
            return Decimal(str(a or 0)) == Decimal(str(b or 0))
        except (InvalidOperation, ValueError):
            pass
    if isinstance(a, (Date, datetime)) or isinstance(b, (Date, datetime)):
        return a == b
    sa, sb = " ".join(str(a).split()), " ".join(str(b).split())
    if sa == sb:
        return True
    # Numeric text: "545" == "545.00", "0" == "0.00"
    try:
        return Decimal(sa) == Decimal(sb)
    except (InvalidOperation, ValueError):
        return False


def _blank(v) -> bool:
    if v is None or v == "":
        return True
    if isinstance(v, Decimal):
        return v == 0
    return False


class Command(BaseCommand):
    help = (
        "Update existing bookings from a Zoho-style Excel export without "
        "wiping any data. Creates only rows that are genuinely absent."
    )

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str)
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Report every change without writing to the database.",
        )
        parser.add_argument(
            "--conflicts", choices=("fill-blanks", "file-wins"),
            default="fill-blanks",
            help=(
                "fill-blanks (default): write only into empty columns. "
                "file-wins: the workbook overwrites a differing stored value."
            ),
        )
        parser.add_argument(
            "--allow-clear", action="store_true",
            help="Let a blank cell erase a stored value (off by default).",
        )
        parser.add_argument(
            "--delegate-number-as", choices=("delegate_count", "delegate_number", "skip"),
            default="delegate_count",
            help=(
                "Where the 'Delegate Number' 0/1 column lands. It holds 0 on "
                "2,422 rows (cancelled and equivalent), which is the countable "
                "flag BookDelegate.delegate_count models, not a sequence number."
            ),
        )
        parser.add_argument(
            "--company-source", choices=("account", "delegate"),
            default="account",
            help=(
                "Which column becomes BookEvent.company_name. The two differ on "
                "35 rows; 'delegate' reproduces what import_booking_excel did."
            ),
        )
        parser.add_argument(
            "--delegate-overrides", choices=("on", "off"), default="on",
            help=(
                "This workbook is delegate-grained: five invoice-level columns "
                "hold DIFFERENT values on different rows of the same invoice "
                "(Payment Status in 58 invoices, Payment Type in 474, Ticket "
                "Tier in 515, Paid/Free in 428, Date Paid in 305). The invoice "
                "can only hold one, so 'on' records the odd rows in the "
                "per-delegate override columns the Bookings table already reads "
                "(delegate_payment_status and friends). 'off' keeps the "
                "invoice's value and loses the distinction."
            ),
        )
        parser.add_argument(
            "--no-create", action="store_true",
            help="Update matched rows only; never insert a new invoice/delegate.",
        )
        parser.add_argument(
            "--report", type=str, default="",
            help="Write a per-field change report to this path.",
        )

    # ── load ──────────────────────────────────────────────────────────────────
    def _load_rows(self, path: Path) -> list[dict]:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [_s(h) for h in next(it)]
        missing = [h for h in REQUIRED_HEADERS if h not in header]
        if missing:
            raise CommandError(
                f"Workbook is missing expected column(s): {', '.join(missing)}"
            )
        rows = []
        for raw in it:
            row = dict(zip(header, raw))
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        wb.close()
        return rows

    def _resolve_event_code(self, raw_code: str) -> Optional[str]:
        """
        The event_code to STORE for a workbook code, or None if it matches no
        Event.

        The workbook writes the code with the edition glued on the end
        ("Feb2027_SFIL-AD27"); every stored row holds the code and the edition in
        separate columns ("Feb2027_SFIL-AD" + 2027), and BookEvent.save() is what
        splits them. So when the code resolves only after the trailing year is
        removed, the RAW code is returned and left for save() to split — return
        the trimmed form instead and the invoice lands with edition NULL, unlike
        its nine siblings.

        The workbook's own spelling is kept in preference to the catalogue's. The
        catalogue holds "FEB2027_SFIL-AD" upper-case while all ten SFIL bookings
        hold "Feb2027_SFIL-AD", and event_code is compared exactly (==) in
        BookEvent.save()'s name lookup and across the reporting queries, so
        importing the catalogue's casing would split one event into two codes.
        """
        code = _s(raw_code)
        if not code:
            return None
        if code.lower() in self._event_by_code:
            return code
        trimmed = re.sub(r"\s*-?\s*\d{2,4}$", "", code).strip()
        if trimmed and trimmed.lower() in self._event_by_code:
            return code
        master, year = normalize_master_code(code), extract_year_from_code(code)
        ev = self._event_by_master.get((master, year))
        return ev.event_code if ev else None

    # ── main ──────────────────────────────────────────────────────────────────
    def handle(self, *args, **opts):
        path = Path(opts["excel_path"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")
        dry = opts["dry_run"]
        policy = opts["conflicts"]
        allow_clear = opts["allow_clear"]
        del_num_target = opts["delegate_number_as"]
        company_col = (
            "Account Company" if opts["company_source"] == "account"
            else "Delegate Company"
        )
        prefix = "[DRY RUN] " if dry else ""

        self.stdout.write(f"Reading {path} ...")
        rows = self._load_rows(path)
        self.stdout.write(f"  {len(rows):,} data rows.")

        # ── lookups ───────────────────────────────────────────────────────────
        self._event_by_code = {}
        self._event_by_master = {}
        for ev in Event.objects.all():
            self._event_by_code.setdefault(ev.event_code.lower(), ev.event_code)
            master = normalize_master_code(ev.event_code)
            self._event_by_master.setdefault((master, extract_year_from_code(ev.event_code)), ev)
            if ev.event_date:
                self._event_by_master.setdefault((master, ev.event_date.year), ev)

        users: dict[str, User] = {}
        for u in User.objects.all():
            full = f"{u.first_name} {u.last_name}".strip().lower()
            if full:
                users.setdefault(full, u)
            users.setdefault(u.username.lower(), u)
            if u.email:
                users.setdefault(u.email.lower(), u)

        invoices = {b.invoice_number: b for b in BookEvent.objects.all()}
        delegates_by_email: dict[tuple, BookDelegate] = {}
        delegates_by_name: dict[tuple, BookDelegate] = {}
        for d in BookDelegate.objects.all():
            delegates_by_email.setdefault((d.invoice_id, (d.email or "").lower()), d)
            delegates_by_name.setdefault((d.invoice_id, _name_key(d.first_name, d.last_name)), d)

        # ── group workbook rows by invoice ────────────────────────────────────
        by_invoice: dict[str, list[dict]] = defaultdict(list)
        no_invoice = 0
        for row in rows:
            inum = _s(row.get("Invoice Number"))
            if not inum:
                no_invoice += 1
                continue
            by_invoice[inum].append(row)
        if no_invoice:
            self.stdout.write(self.style.WARNING(
                f"  {no_invoice} row(s) have no Invoice Number and were skipped."
            ))
        self.stdout.write(f"  {len(by_invoice):,} invoices in the workbook.")

        # ── counters ──────────────────────────────────────────────────────────
        field_changes: dict[str, int] = defaultdict(int)
        field_skipped: dict[str, int] = defaultdict(int)   # differs, policy withheld it
        changed_invoices: list[tuple[BookEvent, dict]] = []
        changed_delegates: list[tuple[BookDelegate, dict]] = []
        new_invoices: list[BookEvent] = []
        new_delegates: list[tuple[BookDelegate, str]] = []
        issues: list[str] = []
        change_log: list[str] = []

        def apply(obj, field, new, label, pending: dict, fill_only=False):
            """
            Stage `new` onto obj.field under the conflict policy.

            fill_only forces fill-blanks for one field whatever --conflicts says.
            It is set for the invoice's contact_* columns, which name ONE of the
            invoice's delegates: the workbook's row order is not the order the
            first import read, so "the first row's delegate" is a different
            person for 459 invoices and file-wins would rename the contact on
            every one of them for no reason.
            """
            old = getattr(obj, field)
            if _same(old, new):
                return
            if _blank(new) and not allow_clear:
                if not _blank(old):
                    field_skipped[label] += 1
                return
            if (policy == "fill-blanks" or fill_only) and not _blank(old):
                field_skipped[label] += 1
                return
            pending[field] = (old, new)

        for inum, irows in by_invoice.items():
            first = irows[0]
            be = invoices.get(inum)
            is_new = be is None

            if is_new:
                if opts["no_create"]:
                    issues.append(f"invoice `{inum}` absent and --no-create set")
                    continue
                code = self._resolve_event_code(first.get("Event Code"))
                if not code:
                    issues.append(
                        f"invoice `{inum}`: event code "
                        f"`{_s(first.get('Event Code'))}` matches no Event — skipped"
                    )
                    continue
                be = BookEvent(invoice_number=inum, event_code=code, source="manual")

            rep_name = _s(first.get("Sales Executive"))
            rep = users.get(rep_name.lower()) if rep_name else None
            if rep_name and not rep:
                issues.append(
                    f"invoice `{inum}`: sales executive `{rep_name}` matches no "
                    f"user — left as stored"
                )

            pending: dict = {}
            apply(be, "payment_status", _s(first.get("Payment Status")) or "Pending",
                  "invoice.payment_status", pending)
            apply(be, "booking_code", _s(first.get("Booking Code")),
                  "invoice.booking_code", pending)
            # BookEvent.save() re-derives event_name from the Event catalogue,
            # but only on an EXACT event_code match, and the catalogue's casing
            # differs from the stored bookings' for some codes — so a new invoice
            # would be left with whatever name it was created with. Every one of
            # the 6,113 stored invoices already agrees with this column, so this
            # is a fill for new rows and a no-op for the rest.
            apply(be, "event_name", _s(first.get("Event Name")),
                  "invoice.event_name", pending)
            apply(be, "request_date", _date(first.get("Request Date")),
                  "invoice.request_date", pending)
            apply(be, "invoice_date", _date(first.get("Invoice Date")),
                  "invoice.invoice_date", pending)
            apply(be, "payment_due_date", _date(first.get("Payment Due")),
                  "invoice.payment_due_date", pending)
            apply(be, "payment_date", _date(first.get("Date Paid")),
                  "invoice.payment_date", pending)
            apply(be, "payment_type", _s(first.get("Payment Type")),
                  "invoice.payment_type", pending)
            apply(be, "paid_or_free", _s(first.get("Paid/Free")),
                  "invoice.paid_or_free", pending)
            apply(be, "ticket_tier", _s(first.get("Ticket Tier")),
                  "invoice.ticket_tier", pending)
            apply(be, "parent_code", _s(first.get("Parent Code")),
                  "invoice.parent_code", pending)
            apply(be, "discount", _dec(first.get("Discount")),
                  "invoice.discount", pending)
            apply(be, "add_ons", _s(first.get("Add-Ons")),
                  "invoice.add_ons", pending)
            apply(be, "reference", _s(first.get("Ref")),
                  "invoice.reference", pending)
            apply(be, "company_name", _s(first.get(company_col)),
                  "invoice.company_name", pending)
            apply(be, "accounts_contact_email", _s(first.get("Accounts Contact")),
                  "invoice.accounts_contact_email", pending)
            apply(be, "contact_name", _s(first.get("Name")),
                  "invoice.contact_name", pending, fill_only=True)
            apply(be, "contact_email", _s(first.get("Delegate Email")),
                  "invoice.contact_email", pending, fill_only=True)
            apply(be, "contact_phone", _phone(first.get("Direct Line")),
                  "invoice.contact_phone", pending, fill_only=True)
            if rep:
                apply(be, "sales_executive", rep, "invoice.sales_executive", pending)
            # delegate_count on the invoice is the number of rows it carries.
            apply(be, "delegate_count", len(irows), "invoice.delegate_count", pending)

            for f, (old, new) in pending.items():
                field_changes[f"invoice.{f}"] += 1
                setattr(be, f, new)
                if len(change_log) < 4000:
                    change_log.append(
                        f"| `{inum}` | invoice.{f} | {old!r} | {new!r} |"
                    )
            if is_new:
                new_invoices.append(be)
                invoices[inum] = be
            elif pending:
                changed_invoices.append((be, pending))

            # ── delegates ─────────────────────────────────────────────────────
            for idx, row in enumerate(irows):
                email = _s(row.get("Delegate Email"))
                full  = _s(row.get("Name"))
                fn, ln = _split_name(full)
                by_email = delegates_by_email.get((inum, email.lower()))
                by_name = delegates_by_name.get((inum, _name_key(fn, ln)))
                # Email first, EXCEPT when it lands on a different person. Two
                # workbook rows share one address on the same invoice (Martin and
                # Anastasha Renaud; Austin Ali and Oz Ruiz), and the stored copy
                # of the second person carries the synthetic
                # `dup-<hash>@import.local` the first import gave them. Matching
                # on the address alone pointed BOTH rows at the first person, so
                # the second person's row was silently never updated and the
                # first person's was written twice.
                if by_name is not None and by_email is not None and by_email is not by_name:
                    bd = by_name
                else:
                    bd = by_email or by_name
                d_new = bd is None
                if d_new:
                    if opts["no_create"]:
                        issues.append(
                            f"delegate `{full}` <{email}> on `{inum}` absent and "
                            f"--no-create set"
                        )
                        continue
                    bd = BookDelegate(
                        invoice_id=inum,
                        event_code=be.event_code,
                        first_name=fn,
                        email=email or f"noemail.{idx}@{inum}.import",
                    )

                dp: dict = {}
                apply(bd, "first_name", fn, "delegate.first_name", dp)
                apply(bd, "last_name", ln, "delegate.last_name", dp)
                apply(bd, "position", _s(row.get("Job Title")),
                      "delegate.position", dp)
                apply(bd, "company_name_raw", _s(row.get("Delegate Company")),
                      "delegate.company_name_raw", dp)
                apply(bd, "phone_number", _phone(row.get("Direct Line")),
                      "delegate.phone_number", dp)
                apply(bd, "booking_code", _s(row.get("Booking Code")),
                      "delegate.booking_code", dp)
                apply(bd, "attendance",
                      "Confirmed" if _bool(row.get("Attendance - IN?")) else "Pending",
                      "delegate.attendance", dp)
                apply(bd, "discount", _dec(row.get("Discount")),
                      "delegate.discount", dp)
                apply(bd, "add_ons", _s(row.get("Add-Ons")),
                      "delegate.add_ons", dp)
                apply(bd, "reference", _s(row.get("Ref")),
                      "delegate.reference", dp)
                # ── per-delegate overrides ────────────────────────────────────
                # Only where this row disagrees with the invoice it shares. A
                # row that agrees leaves the column NULL, which is what the
                # serializers read as "inherit the invoice".
                if opts["delegate_overrides"] == "on":
                    row_vals = {
                        "delegate_payment_status": (_s(row.get("Payment Status")), be.payment_status),
                        "delegate_payment_type":   (_s(row.get("Payment Type")), be.payment_type),
                        "delegate_paid_or_free":   (_s(row.get("Paid/Free")), be.paid_or_free),
                        "delegate_ticket_tier":    (_s(row.get("Ticket Tier")), be.ticket_tier),
                        "delegate_payment_date":   (_date(row.get("Date Paid")), be.payment_date),
                    }
                    for ofield, (rv, iv) in row_vals.items():
                        if _blank(rv) or _same(rv, iv):
                            continue
                        apply(bd, ofield, rv, f"delegate.{ofield}", dp)

                if del_num_target != "skip":
                    raw = row.get("Delegate Number")
                    try:
                        num = int(float(raw)) if raw not in (None, "") else 1
                    except (TypeError, ValueError):
                        num = 1
                    # delegate_count is declared choices=[(0,"0"),(1,"1")].
                    if del_num_target == "delegate_count":
                        num = 1 if num >= 1 else 0
                    apply(bd, del_num_target, num, f"delegate.{del_num_target}", dp)

                for f, (old, new) in dp.items():
                    field_changes[f"delegate.{f}"] += 1
                    setattr(bd, f, new)
                    if len(change_log) < 4000:
                        change_log.append(
                            f"| `{inum}` / {full} | delegate.{f} | {old!r} | {new!r} |"
                        )
                if d_new:
                    # unique_together (invoice, email): two workbook rows share an
                    # email under one invoice, so a genuinely new second person
                    # needs an address of their own.
                    key = (inum, (bd.email or "").lower())
                    if key in delegates_by_email:
                        local, _, domain = bd.email.partition("@")
                        bd.email = f"{local}.{idx}@{domain or 'import.local'}"
                        issues.append(
                            f"delegate `{full}` on `{inum}` shares an email with "
                            f"another row; stored as {bd.email}"
                        )
                    delegates_by_email[(inum, (bd.email or "").lower())] = bd
                    delegates_by_name[(inum, _name_key(bd.first_name, bd.last_name))] = bd
                    new_delegates.append((bd, full))
                elif dp:
                    changed_delegates.append((bd, dp))

        # ── report ────────────────────────────────────────────────────────────
        self.stdout.write("")
        self.stdout.write(f"{prefix}Policy         : {policy}"
                          f"{' (+allow-clear)' if allow_clear else ''}")
        self.stdout.write(f"{prefix}company_name   : {company_col}")
        self.stdout.write(f"{prefix}Delegate Number: {del_num_target}")
        self.stdout.write("")
        self.stdout.write(f"{prefix}New invoices   : {len(new_invoices):,}")
        self.stdout.write(f"{prefix}New delegates  : {len(new_delegates):,}")
        self.stdout.write(f"{prefix}Invoices changed : {len(changed_invoices):,}")
        self.stdout.write(f"{prefix}Delegates changed: {len(changed_delegates):,}")
        if issues:
            self.stdout.write(self.style.WARNING(f"{prefix}Issues         : {len(issues)}"))

        if field_changes:
            self.stdout.write("\nWrites by field:")
            width = max(len(k) for k in field_changes)
            for f, n in sorted(field_changes.items(), key=lambda kv: -kv[1]):
                self.stdout.write(f"  {f:<{width}}  {n:>6,}")
        if field_skipped:
            self.stdout.write("\nDiffers but withheld by the policy:")
            width = max(len(k) for k in field_skipped)
            for f, n in sorted(field_skipped.items(), key=lambda kv: -kv[1]):
                self.stdout.write(f"  {f:<{width}}  {n:>6,}")

        if opts["report"]:
            rp = Path(opts["report"])
            lines = [
                "# Remaining-data import report\n\n",
                f"Workbook: `{path}`  \n",
                f"Policy: **{policy}**"
                f"{' + allow-clear' if allow_clear else ''}  \n",
                f"company_name from: **{company_col}**  \n",
                f"Delegate Number -> **{del_num_target}**  \n\n",
                f"- New invoices: {len(new_invoices)}\n",
                f"- New delegates: {len(new_delegates)}\n",
                f"- Invoices updated: {len(changed_invoices)}\n",
                f"- Delegates updated: {len(changed_delegates)}\n\n",
                "## Writes by field\n\n| Field | Rows |\n|---|---|\n",
            ]
            for f, n in sorted(field_changes.items(), key=lambda kv: -kv[1]):
                lines.append(f"| {f} | {n} |\n")
            if field_skipped:
                lines.append("\n## Differs but withheld by the policy\n\n"
                             "| Field | Rows |\n|---|---|\n")
                for f, n in sorted(field_skipped.items(), key=lambda kv: -kv[1]):
                    lines.append(f"| {f} | {n} |\n")
            if issues:
                lines.append(f"\n## Issues ({len(issues)})\n\n")
                for i in issues:
                    lines.append(f"- {i}\n")
            lines.append(f"\n## Changes ({len(change_log)} shown)\n\n"
                         "| Row | Field | Stored | Workbook |\n|---|---|---|---|\n")
            lines.extend(f"{l}\n" for l in change_log)
            rp.write_text("".join(lines), encoding="utf-8")
            self.stdout.write(f"\nReport -> {rp}")

        if dry:
            self.stdout.write(self.style.SUCCESS("\nDry run complete — nothing written."))
            return

        # ── write ─────────────────────────────────────────────────────────────
        # save() per row on purpose: BookEvent.save()/BookDelegate.save() are the
        # chokepoints that canonicalise booking_code, split edition out of
        # event_code, re-derive event_name and booked_on, and fill a blank
        # accounts contact. bulk_create/bulk_update bypass every one of them,
        # which is how the wiping importer has to re-implement them by hand.
        self.stdout.write("\nWriting ...")
        with transaction.atomic():
            for be in new_invoices:
                be.save()
            for be, _ in changed_invoices:
                be.save()
            self.stdout.write(
                f"  invoices: {len(new_invoices):,} created, "
                f"{len(changed_invoices):,} updated."
            )
            for bd, _ in new_delegates:
                bd.save()
            for bd, _ in changed_delegates:
                bd.save()
            self.stdout.write(
                f"  delegates: {len(new_delegates):,} created, "
                f"{len(changed_delegates):,} updated."
            )

        self.stdout.write(self.style.SUCCESS(
            f"\nDone. {sum(field_changes.values()):,} field write(s) across "
            f"{len(changed_invoices) + len(new_invoices):,} invoices and "
            f"{len(changed_delegates) + len(new_delegates):,} delegates."
        ))
